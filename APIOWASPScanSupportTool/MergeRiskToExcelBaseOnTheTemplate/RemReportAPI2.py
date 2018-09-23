#encoding=utf-8
from bs4 import BeautifulSoup
import os
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook import Workbook
import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,colors,borders
import collections
import re
import pickle
import sys
import shutil
import logging
#to avoid the issue UnicodeEncodeError: 'ascii' codec can't encode character
reload(sys)
sys.setdefaultencoding('utf8')

''' It's to generate the report automatically for the remediation report'''

class remReport():
	'''
	self.BuildVersion for the build version
	self.SumTitleLine for the summary sheet title
	self.OtherTitleLine for the sheet title exclude summary
	self.CurrentLocation for the base folder
	self.ReportFileName for the report file name
	self.DictDataFileName for the saved all the risk dict data file name
	self.DictURLFileName for the saved url dict data file name
	self.DictSectionFileName for the saved section dict data file name
	self.DictDescriptionFileName for the saved description dict data file name
	self.DictHyperFileName for the saved summary hyper name dict data file name
	self.DictFileList=for the saved dict flie list
	self.MergeFolderName for the merged root folder name
	self.BackupFolderName for the back up folder name
	self.Phase for the build phase
	self.Dict for the risk dict data
	self.DictMerge for the new merge dict data
	self.DictURL for the url dict data
	self.DictDescription for the description dict data
	self.DictSection for the section dict data
	self.DictHyper for the saved summary hyper name dict dat
	self.SummarySheetName for the summary sheet name
	self.DataSplit for the specific char split
	self.ModuleFolderList for the module folder name list
	
	
	self.cmdFlag for the print information in the console flag automatically, default is True, means print the information in the console.
	self.moveFlag for the move merge html flag, default is None, mean not to move the html folder to the related one.
	self.mergeFlag for the merge judge flag, default is none, means to judge merge action 
	'''

	def __init__(self,BuildVersion,Phase,mergeFlag=None,moveFlag=None,cmdFlag=True):
		self.SumTitleLine=['Finding','Severity','Status','Description','ISO RESPONSE','Jira Ticket','Patch ETA','Comments','Affected URLs','Attack','Evidence','Suggested Resolution']
		self.SumSubTitleLine=["","Add the Severity Finding","Pick a Status from below(See Status Key Description Tab)","Provide issue description here","PLEASE REVIEW AND PROVIDE ADDITIONAL INPUT ABOUT THE OBJECTIONS","Insert Jira ticket number here","Provide patch ETA and product version number","Insert comments here","Provide affected URLs here","Describe issue/attack here","Provide code information or web page screen shot as evidence","Provide suggested resolution steps."]
		self.OtherTitleLine=["Description","Section","URL","Note"]
		self.ModelRiskTitleLine=["RiskLevel","Number of Alerts"]
		self.CurrentLocation=os.getcwd()
		self.ReportFileName="DEV-CODE-REMED Development Code Remediation Plan for "+BuildVersion+" "+Phase+" OWASP ZAP Scan__FINAL.xlsx"
		self.TemplateFileName="Template.xlsx"
		self.DictDataFileName="Dict.txt"
		self.DictURLFileName="DictURL.txt"
		self.DictSectionFileName="DictSection.txt"
		self.DictDescriptionFileName="DictDescription.txt"
		self.DictHyperFileName="DictHyper.txt"
		self.DictFileList=[self.DictDataFileName,self.DictURLFileName,self.DictSectionFileName,self.DictDescriptionFileName,self.DictHyperFileName]
		self.MergeFolderName="Merge"
		self.BackupFolderName=Phase+"Backup"
		self.BuildVersion=BuildVersion
		self.Phase=Phase
		self.mergeFlag=mergeFlag
		self.Dict=None
		self.DictMerge=None
		self.DictURL=None
		self.DictDescription=None
		self.DictSection=None
		self.DictHyper=None
		self.SummarySheetName="Remediation Plan Template"
		self.DataSplit="||||"
		self.ModuleFolderList=['CAA','CLAPI','CLRESTWS','SITEDOC','WEBSERVICES','CLWIZARD']
		self.cmdFlag=cmdFlag
		self.initLogger()
		self.moveFlag=moveFlag
		self.SummaryColumnWidth=[12,21,31,60,60,29,24,36,89,91,110,37]
		
	def initLogger(self):
		'''initial the logger'''
		os.system("type nul>log.txt")
		logger = logging.getLogger(__name__)
		logger.setLevel(level = logging.INFO)
		self.logger=logger
		
	def logSteps(self,content):
		'''judge whether need to print in the console, and logger the records in the log file'''
		if self.cmdFlag:
			print content
		if len(self.logger.handlers)==0:
			handler = logging.FileHandler("log.txt")
			self.logger.addHandler(handler)
		self.logger.info(content)

	def use_logging(func):
		def wrapper(*args, **kwargs):
			print func.__doc__
			logger = logging.getLogger(__name__)
			logger.setLevel(level = logging.INFO)
			#below line to avoid the duplidate log informaiton
			if len(logger.handlers)==0:
				handler = logging.FileHandler("log.txt")
				logger.addHandler(handler)
			logger.info(func.__doc__)
			return func(*args, **kwargs)
		return wrapper
	
	def commonFormatSetting(self,sheet,cIndex,rIndex):
		'''do the common excel FormatSetting'''
		sheet.cell(column=cIndex,row=rIndex).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
		if rIndex>2 and sheet.title==self.SummarySheetName:
			sheet.cell(column=cIndex,row=rIndex).font=Font(size=16)
		
	def SummarySheetFormatSetting(self,sheet,cIndex,rIndex):
		'''do the summary sheet format'''
		sheet.cell(column=cIndex,row=rIndex).alignment=Alignment(wrap_text = True,horizontal='left',vertical='top')
		if rIndex>2:
			if cIndex==2:
				
				if re.search("High",sheet.cell(column=cIndex,row=rIndex).value):
					sheet.cell(column=cIndex,row=rIndex).fill=PatternFill(fill_type='solid',fgColor=colors.RED)
				elif re.search("Medium",sheet.cell(column=cIndex,row=rIndex).value):
					sheet.cell(column=cIndex,row=rIndex).fill=PatternFill(fill_type='solid',fgColor='00FFB90F')
				elif re.search("Low",sheet.cell(column=cIndex,row=rIndex).value):
					sheet.cell(column=cIndex,row=rIndex).fill=PatternFill(fill_type='solid',fgColor='00CDBE70')
				elif re.search("Informational",sheet.cell(column=cIndex,row=rIndex).value):
					sheet.cell(column=cIndex,row=rIndex).fill=PatternFill(fill_type='solid',fgColor='001E90FF')

				
	def OtherSheetFormatSetting(self,sheet,cIndex,rIndex,RedFlag):
		'''do the sheet exclude summary format '''
		sheet.cell(column=cIndex,row=rIndex).alignment=Alignment(wrap_text = True,horizontal='left',vertical='top')
		if rIndex==1:
			sheet.cell(column=cIndex,row=1).font=Font(size=11,bold=True)
		if re.match(r"High|Medium|Low|Informational",str(sheet.cell(column=cIndex,row=rIndex).value)) or RedFlag:
			sheet.cell(column=cIndex,row=rIndex).font=Font(color=colors.WHITE)
			sheet.cell(column=cIndex,row=rIndex).fill = PatternFill(fill_type='solid',fgColor=colors.RED)
			if cIndex==1:
				RedFlag=True
			else:
				RedFlag=False
			return RedFlag
		elif rIndex>3 and cIndex<=2 and sheet.cell(column=cIndex,row=rIndex).value!="":
			sheet.cell(column=cIndex,row=rIndex).fill = PatternFill(fill_type='solid',fgColor='00FFF8DC')
		return RedFlag
	
	def theModelSheetSummaryReset(self,sheet):
		'''do the model sheet summary section format'''
		if sheet.title in self.ModuleFolderList:
			for rIndex in range(2,6):
				for cIndex in range(1,3):
					sheet.cell(column=cIndex,row=rIndex).font=Font(color=colors.BLACK)
					sheet.cell(column=cIndex,row=rIndex).fill = PatternFill(fill_type='solid',fgColor='00D3D3D3')
					sheet.cell(column=cIndex,row=rIndex).font=Font(size=11,bold=True)

	def getWidth(self,sheet,column_widths,cIndex,rIndex):
		'''get the each column max width'''
		#as some values as HYPERLINK value, so need to check relately.
		cellValue=len(str(sheet.cell(column=cIndex,row=rIndex).value))
		if len(column_widths) >(cIndex-1):
			if cellValue> column_widths[cIndex-1]:
				column_widths[cIndex-1] = cellValue
		else:
			#as the column width isn't adpated relately.
			column_widths += [(cellValue+2)*1.5]
		return column_widths

	def setWidth(self,sheet,column_widths):
		'''reset the column width for the related width'''
		Dict=sheet.column_dimensions
		for i, column_width in enumerate(column_widths):
			if column_width>100:
				column_width=100
			Dict[get_column_letter(i+1)].width =column_width

	def setSummarySheetHeight(self,sheet):
		'''reset the row height'''
		if sheet.title==self.SummarySheetName:
			Dict=sheet.row_dimensions
			MaxRow=sheet.max_row
			for i in range(1,MaxRow+1):
				if i==1:
					Dict[i].height=20
				else:
					Dict[i].height=60
	
	def mergeFlagJudge(self):
		'''if the merge flag, could use the method to judge whether need to merge'''
		if os.path.exists(os.path.join(self.CurrentLocation,self.MergeFolderName)) and len(self.getFileList(os.path.join(self.CurrentLocation,self.MergeFolderName)))!=0:
			mergeFlag=True
		else:
			mergeFlag=False
		return mergeFlag
		
	def splitSpecialWord(self,content):
		'''to split up the word '''
		NewContent=""
		index=0
		for i in content:
			if index>=2 and i!="|":
				NewContent+=i
			elif index<2:
				NewContent+=i
				if i=="|":
					index+=1
		return NewContent.split(self.DataSplit)
	
	def dictDataInitial(self,FileName):
		'''to get the data for each Dict'''
		if self.mergeFlag:
			fr=open(os.path.join(self.CurrentLocation,self.Phase,FileName),'rb')
			data=pickle.load(fr)
			fr.close()
			return data
		else:
			return collections.OrderedDict()
			
	def writeDict(self,FileName,Data):
		'''to save the dict data to the file'''
		fw=open(os.path.join(self.CurrentLocation,self.Phase,FileName),'wb')
		pickle.dump(Data,fw)
		fw.close()
		
	def getFolderList(self,location):
		'''to get the folder list'''
		DirList=[]
		for dirItem in os.listdir(location):
			os.chdir(location)
			if os.path.isdir(dirItem):
				DirList.append(dirItem)
		return DirList
		
	def getFileList(self,location):
		'''to get the file list'''
		FileList=[]
		for root,dir,files in os.walk(location):
			FileList.extend(files)
		return FileList
		
	def splitServerity(self,List):
		'''to split each risk as one list'''
		List1=[]
		index=-1
		for i in List:
			if re.match(r"High|Low|Medium|Informational",i):
				List1+=[[i]]
				index+=1
			else:
				List1[index].append(i)
		return List1
		
	def removeMergeDuplicateData(self):
		'''to remove the duplicated data if the merge data has already exist'''
		for MergeRiskName in self.DictMerge.keys():
			for DictRiskitem in self.Dict.keys():
				if MergeRiskName==DictRiskitem:
					MergeItem=self.splitServerity(self.DictMerge[MergeRiskName])
					DictItem=self.splitServerity(self.Dict[MergeRiskName])
					MergeItem=filter(lambda x:DictItem.count(x)==0,MergeItem[:])
					TempList=[]
					[TempList.extend(item) for item in MergeItem]
					if TempList!=[]:
						self.DictMerge[MergeRiskName]=TempList
					else:
						del self.DictMerge[MergeRiskName]

						
	def removeDuplicateDictData(self):
		'''to remove the duplicated data if the orginal Dict data has already exist'''
		for DictRiskitem in self.Dict.keys():
			DictItem=self.splitServerity(self.Dict[DictRiskitem])
			TempList=[]
			TempList1=[]
			[TempList.append(i) for i in DictItem if i not in TempList]
			[TempList1.extend(i) for i in TempList]
			self.Dict[DictRiskitem]=TempList1
		
		
	def getMergeCommonDict(self,RiskName,Dict,DictMerge,DirItem,Tag,Name):
		'''to get the data for merge Dict or common Dict '''
		if self.mergeFlag:
			if DictMerge.has_key(DirItem+" "+RiskName):
				DictMerge[DirItem+" "+RiskName].append(Tag+self.DataSplit+Name)
			else:
				DictMerge[DirItem+" "+RiskName]=[Tag+self.DataSplit+Name]
		else:
			if Dict.has_key(DirItem+" "+RiskName):
				Dict[DirItem+" "+RiskName].append(Tag+self.DataSplit+Name)
			else:
				Dict[DirItem+" "+RiskName]=[Tag+self.DataSplit+Name]
		return Dict,DictMerge


	def initalDir(self):
		'''initial the module folder for the report html'''
		InitiaModulelFlag=False
		for i in self.ModuleFolderList:
			if not os.path.exists(os.path.join(self.CurrentLocation,i)):
				os.mkdir(os.path.join(self.CurrentLocation,i))
				InitiaModulelFlag=True
		#create backup folder
		if not os.path.exists(os.path.join(self.CurrentLocation,self.BackupFolderName)):
			os.mkdir(os.path.join(self.CurrentLocation,self.BackupFolderName))
		#create merger folder
		if not os.path.exists(os.path.join(self.CurrentLocation,self.MergeFolderName)):
			os.mkdir(os.path.join(self.CurrentLocation,self.MergeFolderName))
		#create merge subfolder
		for i in self.ModuleFolderList:
			if not os.path.exists(os.path.join(self.CurrentLocation,self.MergeFolderName,i)):
				os.mkdir(os.path.join(self.CurrentLocation,self.MergeFolderName,i))
		#create Dict Data
		if not os.path.exists(os.path.join(self.CurrentLocation,self.Phase)):
			os.mkdir(os.path.join(self.CurrentLocation,self.Phase))
			os.chdir(os.path.join(self.CurrentLocation,self.Phase))
			for i in self.DictFileList:
				os.system("type nul>%s" %i)
		if InitiaModulelFlag:
			content="\nInfo:The folder has already been initialed, please put the related html to the related module folder, and rerun to generate the report"
			self.logSteps(content)
			sys.exit(1)
		
	@use_logging
	def findDir(self):
		'''start to search the dir'''
		self.initalDir()
		if self.mergeFlag==None or self.mergeFlag=="":
			self.mergeFlag=self.mergeFlagJudge()
		elif self.mergeFlag==True and len(self.getFileList(os.path.join(self.CurrentLocation,self.MergeFolderName)))==0:
			content="\nError:The merged data hasn't been input, please put the merged report files"
			self.logSteps(content)
			sys.exit(1)
		DirList=self.ModuleFolderList
		return DirList


	@use_logging
	def parseData(self,DirList):
		'''start the parseData'''
		Dict=self.dictDataInitial(self.DictDataFileName)
		DictMerge=collections.OrderedDict()
		DictURL=self.dictDataInitial(self.DictURLFileName)
		DictDescription=self.dictDataInitial(self.DictDescriptionFileName)
		DictSection=self.dictDataInitial(self.DictSectionFileName)
		DictHyper=self.dictDataInitial(self.DictHyperFileName)
		if self.mergeFlag:
			Location=os.path.join(self.CurrentLocation,self.MergeFolderName)
		else:
			Location=os.path.join(self.CurrentLocation)
		for DirItem in DirList:
			os.chdir(os.path.join(Location,DirItem))
			for root, dirs, files in os.walk(os.path.join(Location,DirItem)):
				if dirs==[]:
					os.chdir(os.path.join(root))
					for file in files:
						bare_name =  os.path.splitext(file)[0]
						extension_name = os.path.splitext(file)[1]
						if extension_name == '.html':
						#it needs to use the html.parser, otherwise it has the warpper warning.
							soup = BeautifulSoup(open(bare_name+extension_name),"html.parser")
							for table in soup.find_all('table'):
								RiskLevel=""
								RiskName=""
								for tr in table.find_all('tr'):
									if "of Alerts" not in table.getText():
										Thcells = tr.find_all('th')
										if len(Thcells)>=2:
											RiskLevel=Thcells[0].find(text=True)
											RiskName=Thcells[1].find(text=True)
											Dict,DictMerge=self.getMergeCommonDict(RiskName,Dict,DictMerge,DirItem,RiskLevel,RiskName)
										Tdcells = tr.find_all('td')
										if len(Tdcells)>=2:
											Tag=Tdcells[0].getText().strip().replace(r"\xa0","")
											Name=str(Tdcells[1].getText().strip().replace(r"\xa0",""))
											if Name=="":
												Name="None"
											if RiskName=="" and RiskName=="":
												RiskLevel=Tag
												RiskName=Name
											Dict,DictMerge=self.getMergeCommonDict(RiskName,Dict,DictMerge,DirItem,Tag,Name)
											if Tag=="URL" and DictURL.has_key(DirItem+" "+RiskName):
												DictURL[DirItem+" "+RiskName].append(Name)
											elif Tag=="URL":
												DictURL[DirItem+" "+RiskName]=[Name]
											if Tag=="Description" and DictDescription.has_key(DirItem+" "+RiskName):
												pass
											elif Tag=="Description":
												DictDescription[DirItem+" "+RiskName]=Name
										if DictSection.has_key(DirItem+" "+RiskName) and bare_name not in DictSection[DirItem+" "+RiskName]:
											DictSection[DirItem+" "+RiskName].append(bare_name)
										elif not DictSection.has_key(DirItem+" "+RiskName):
											DictSection[DirItem+" "+RiskName]=[bare_name]
		self.Dict=Dict
		self.removeDuplicateDictData()
		self.DictMerge=DictMerge
		self.removeMergeDuplicateData()
		self.DictURL=DictURL
		self.DictDescription=DictDescription
		self.DictSection=DictSection
		self.DictHyper=DictHyper


	def initialWorkBook(self):
		'''initial the workbook'''
		if self.mergeFlag:
			if len(self.DictMerge.keys())==0:
				content="\nWarning:The merged data is duplicated, no need to update"
				self.logSteps(content)
				sys.exit(1)
			else:
				try:
					wb=load_workbook(os.path.join(self.CurrentLocation,self.ReportFileName))
					content="start to merge the data"
					self.logSteps(content)
				except:
					content="\nWarning:Sorry,the report excel is missing, you couldn't merge the data,\nplease clear the merge folder at first to initial new report, then you could merge"
					self.logSteps(content)
					sys.exit(1)
		else:
			if len(self.Dict.keys())==0:
				content="\nWarning:There is no report to initial"
				self.logSteps(content)
				sys.exit(1)
			wb=load_workbook(os.path.join(self.CurrentLocation,self.TemplateFileName))
			content="start to initial new report"
			self.logSteps(content)
		return wb

	def SummaryHyperNewSheet(self,workbook,summarySheet,rowNum,modelRiskName):
		'''set the summary hyperlink and create related sheet '''
		self.DictHyper[modelRiskName]=str(rowNum-2)
		# summarySheet['A%s' %(rowNum)].value =rowNum-2



	@use_logging
	def SummaryCreate(self,workbook):
		'''start to update the summary data'''
		self.DictModel=collections.OrderedDict()
		sheet=workbook[self.SummarySheetName]
		if self.mergeFlag:
			for item in self.DictMerge.keys():
				if self.DictHyper.has_key(item):
					sheet.cell(column=4,row=int(self.DictHyper[item])+2).value=item+" is hit when "+",".join(self.DictSection[item])
				else:
					MaxRow=sheet.max_row
					self.summaryContentInput(MaxRow-1,item,workbook,sheet,self.DictMerge)
		else:
			for index,rowValue in enumerate(["SumTitleLine"]+self.Dict.keys()):
				if index>0:
					self.summaryContentInput(index,rowValue,workbook,sheet,self.Dict)


	def summaryContentInput(self,index,rowValue,workbook,sheet,sheetDict):
		'''summary content input'''
		rowContent=[str(index),sheetDict[rowValue][0].split()[0],"",rowValue+" is hit when "+",".join(self.DictSection[rowValue]),"","",self.BuildVersion,"","\n".join(list(set(self.DictURL[rowValue]))),self.DictDescription[rowValue],"",re.search(r"Solution(.*)","".join(sheetDict[rowValue])).group(1).split(self.DataSplit)[1]]
		sheet.append(rowContent)
		self.SummaryHyperNewSheet(workbook,sheet,index+2,rowValue)

	def riskContentInput(self,rowNum,sheet,riskContent):
		rowNum+=1
		if re.match(r"High|Medium|Low|Informational",riskContent):
			sheet.cell(column=1,row=rowNum).value=""
			sheet.cell(column=2,row=rowNum).value=""
			rowNum+=1
			sheet.cell(column=1,row=rowNum).value=riskContent.split(self.DataSplit)[0]
			sheet.cell(column=2,row=rowNum).value=riskContent.split(self.DataSplit)[1]
			
		else:
			sheet.cell(column=1,row=rowNum).value=riskContent.split(self.DataSplit)[0]
			sheet.cell(column=2,row=rowNum).value=riskContent.split(self.DataSplit)[1]
		return rowNum



	@use_logging
	def FormatSheetData(self,workbook):
		'''start to format all of the sheets'''
		sheetNames=workbook.sheetnames
		for sheetName in sheetNames:
			
			sheet=workbook[sheetName]
			MaxRow=sheet.max_row
			column_widths=[]
			if sheetName==self.SummarySheetName:
				TitleLine=self.SumTitleLine
			elif sheetName in self.ModuleFolderList:
				TitleLine=self.ModelRiskTitleLine
			else:
				TitleLine=self.OtherTitleLine
			for rIndex in range(1,MaxRow+1):
				if sheetName!=self.SummarySheetName:
					RedFlag=False
				for cIndex in range(1,len(TitleLine)+1):
					#The Fone will be replaced automatically, as the default is black. so need to set the red in the below
					self.commonFormatSetting(sheet,cIndex,rIndex)
					if sheetName!=self.SummarySheetName:
						RedFlag=self.OtherSheetFormatSetting(sheet,cIndex,rIndex,RedFlag)
					else:
						self.SummarySheetFormatSetting(sheet,cIndex,rIndex)
					column_widths=self.getWidth(sheet,column_widths,cIndex,rIndex)
			if sheetName==self.SummarySheetName:
				column_widths=self.SummaryColumnWidth
			self.theModelSheetSummaryReset(sheet)
			#set the width
			self.setWidth(sheet,column_widths)
			self.setSummarySheetHeight(sheet)
		
	def mergedDict(self):
		'''merge the new dict to the exist common dict'''
		if self.mergeFlag:
			for MergeRiskName in self.DictMerge.keys():
				if self.Dict.has_key(MergeRiskName):
					self.Dict[MergeRiskName].extend(self.DictMerge[MergeRiskName])
				else:
					self.Dict[MergeRiskName]=self.DictMerge[MergeRiskName]
	
	def backupFiles(self):
		'''back up the files'''
		for root, dirs, files in os.walk(os.path.join(self.CurrentLocation,self.Phase)):
			for file in files:
				try:
					os.remove(os.path.join(self.CurrentLocation,self.BackupFolderName,file))
				except WindowsError:
					pass
				shutil.copy(os.path.join(self.CurrentLocation,self.Phase,file),os.path.join(self.CurrentLocation,self.BackupFolderName))
				content="\nWarning:backup the file %s to the backup folder %s" %(file,self.BackupFolderName)
				report.logSteps(content)
		try:
			os.remove(os.path.join(self.CurrentLocation,self.BackupFolderName,self.ReportFileName))
		except WindowsError:
			pass
		try:
			shutil.copy(os.path.join(self.CurrentLocation,self.ReportFileName),os.path.join(self.CurrentLocation,self.BackupFolderName))
			content="\nWarning:backup the report file %s to the backup folder %s" %(file,self.BackupFolderName)
			report.logSteps(content)
		except IOError:
			pass
		
		
	def moveMergedHtml(self,DirList):
		'''move the html to the merge folder'''
		if self.moveFlag:
			content="start to move the merged html to target folder"
			report.logSteps(content)
			if not self.mergeFlag:
				Location=os.path.join(self.CurrentLocation)
				
			else:
				Location=os.path.join(self.CurrentLocation,self.MergeFolderName)
				for DirItem in DirList:
					for root, dirs, files in os.walk(os.path.join(Location,DirItem)):
						os.chdir(os.path.join(Location,DirItem))
						for file in files:
							try:
								os.remove(os.path.join(self.CurrentLocation,self.BackupFolderName,file))
							except WindowsError:
								pass
							try:
								shutil.copy(os.path.join(self.CurrentLocation,DirItem,file),os.path.join(self.CurrentLocation,self.BackupFolderName))
								content="\nWarning:the duplicated file %s has been moved to backup folder %s" %(file,self.BackupFolderName)
								report.logSteps(content)
							except IOError:
								pass
							try:
								os.remove(os.path.join(self.CurrentLocation,DirItem,file))
							except WindowsError:
								pass
							shutil.copy(file,os.path.join(self.CurrentLocation,DirItem,file))
							os.remove(file)
							content="\nWarning:move the file %s to the folder %s" %(file,os.path.join(self.CurrentLocation,DirItem))
							report.logSteps(content)


							
	@use_logging
	def mergeModelValue(self,workbook):
		'''merge the related risk into the related model sheet'''
		self.mergedDict()
		sheetNames=workbook.sheetnames
		DictTotalNum={}
		#to account the specific sheet number, needs the dict, and every key needs to put the specific value
		for item in self.ModuleFolderList:
			if item in sheetNames:
				workbook.remove(workbook[str(item)])
			DictTotalNum[item]={"HighNum":0,"MiddleNum":0,"LowNum":0,"InformationalNum":0}
		with open("2.txt","w") as fp:
			fp.write(str(self.Dict))
		for item in self.Dict.keys():
			SheetName=item.split()[0]
			sheetNames=workbook.sheetnames
			if SheetName not in sheetNames:
				workbook.create_sheet(SheetName)
			sheet=workbook[SheetName]
			rowNum=sheet.max_row
			rowNum+=4
			for i in self.Dict[item]:
				rowNum=self.riskContentInput(rowNum,sheet,i)
				if re.match(r"High",i):
					DictTotalNum[SheetName]["HighNum"]+=1
				elif re.match(r"Medium",i):
					DictTotalNum[SheetName]["MiddleNum"]+=1
				elif re.match(r"Low",i):
					DictTotalNum[SheetName]["LowNum"]+=1
				elif re.match(r"Informational",i):
					DictTotalNum[SheetName]["InformationalNum"]+=1
			sheet.cell(column=1,row=1).value="RiskLevel"
			sheet.cell(column=2,row=1).value="Number of Alerts"
			sheet.cell(column=1,row=2).value="High"
			sheet.cell(column=2,row=2).value=DictTotalNum[SheetName]["HighNum"]
			sheet.cell(column=1,row=3).value="Medium"
			sheet.cell(column=2,row=3).value=DictTotalNum[SheetName]["MiddleNum"]
			sheet.cell(column=1,row=4).value="Low"
			sheet.cell(column=2,row=4).value=DictTotalNum[SheetName]["LowNum"]
			sheet.cell(column=1,row=5).value="Informational"
			sheet.cell(column=2,row=5).value=DictTotalNum[SheetName]["InformationalNum"]
		
	@use_logging
	def saveExcel(self,workbook):
		'''Back up the orginal dict and report files,save the report excel, then merge the merged data to the risk dict data, and finial update all of the dict files'''
		self.backupFiles()
		workbook.save(os.path.join(self.CurrentLocation,self.ReportFileName))
		self.writeDict(self.DictDataFileName,self.Dict)
		self.writeDict(self.DictURLFileName,self.DictURL)
		self.writeDict(self.DictSectionFileName,self.DictSection)
		self.writeDict(self.DictDescriptionFileName,self.DictDescription)
		self.writeDict(self.DictHyperFileName,self.DictHyper)
		
if __name__ == "__main__":
	#__init__(self,BuildVersion,Phase,mergeFlag=None,moveFlag=None,cmdFlag=True):
	report=remReport("CIS622","RC5",False)
	DirList=report.findDir()
	report.parseData(DirList)
	wb=report.initialWorkBook()
	report.SummaryCreate(wb)
	report.mergeModelValue(wb)
	report.FormatSheetData(wb)
	report.saveExcel(wb)
	report.moveMergedHtml(DirList)
	content="\nPlease go to the below location %s to check the report" %(os.path.join(report.CurrentLocation,report.ReportFileName))
	report.logSteps(content)











