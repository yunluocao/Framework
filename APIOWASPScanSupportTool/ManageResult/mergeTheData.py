#encoding=utf-8
import os
import json
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def MergeFunction(Owner):
	with open("FileList.txt","r") as fp:
		jsObjList=json.loads(fp.read())
	wb=load_workbook("OWASP_Scan_Staus.xlsx")
	sheet=wb["Summary"]
	for item in jsObjList:
		for rowIndex in range(1,sheet.max_row+1):
			if item==str(sheet.cell(column=2,row=rowIndex+1).value).strip():
				sheet.cell(column=3,row=rowIndex+1).value="Pass"
				sheet.cell(column=4,row=rowIndex+1).value=Owner
	wb.save("OWASP_Scan_Staus.xlsx")
	
MergeFunction("Jonson")