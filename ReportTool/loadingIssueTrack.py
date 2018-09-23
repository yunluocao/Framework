#encoding=utf-8
import re
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,colors,borders
import sys
import os
import traceback

class generateSummary(object):
	
	def __init__(self,propertyFile):
		self.properties=self.parseProp(propertyFile)
		self.excelName=self.properties['excelName']
		self.wb=load_workbook(self.excelName)
		self.sheetTitle=eval(self.properties['sheetTitle'])
		self.columnWidth=eval(self.properties['columnWidth'])
		try:
			self.sheet=self.wb[self.properties['sheetName']]
		except KeyError as e:
			self.sheet=self.createSheet(self.properties['sheetName'],self.sheetTitle)
		self.maxRow=self.sheet.max_row
		self.DupList=[]
		self.rowid=2
		self.platformName=self.properties['platformName']
		self.platformDict=eval(self.properties['platformDict'])
		self.suitePath=self.properties['suitePath']
		self.logFile=self.properties['logFile']
		self.fontName=self.properties['fontName']
		self.fontSize=self.properties['fontSize']
		self.printDuplicatedFlag=eval(self.properties['printDuplicatedFlag'])
		
		
	def parseProp(self,propertyFile):
		properties={}
		with open(propertyFile,"r") as fp:
			for line in fp:
				line=line.strip()
				if line.find("=")>0 and not line.startswith("#"):
					strs=line.split("=")
					properties[strs[0].strip()]=strs[1].strip()
		return properties
	
	def createSheet(self,sheetName,title):
		sheet=self.wb.create_sheet(sheetName)
		sheet.append(title)
		self.sheetFormatSetting(sheet,title)
		sheet.merge_cells('C1:D1')
		return sheet
		
	def sheetFormatSetting(self,sheet,title):
		for cIndex in range(1,len(title)+1):
			sheet.cell(column=cIndex,row=1).font=Font(size=8,bold=True,name='Arial')
			'00FFB90F'
			sheet.cell(column=cIndex,row=1).fill=PatternFill(fill_type='solid',fgColor='007EC0EE')
			sheet.cell(column=cIndex,row=1).alignment=Alignment(wrap_text = True,horizontal='center',vertical='center')
			sheet.cell(column=cIndex,row=1).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
		print dir(sheet.cell(column=1,row=1))
		Dict=sheet.column_dimensions
		for i, column_width in enumerate(self.columnWidth):
			Dict[get_column_letter(i+1)].width =column_width
		Dict=sheet.row_dimensions
		Dict[1].height=26.5
		
		
	def read_the_summaryValue(self):
		contentList=[]
		with open(self.logFile,"r") as fp:
			for i in fp:
				if i.startswith("unidentified") or i.startswith("unknown"):
					element1=re.search("%s/(.*) - (.*)"%self.suitePath,i).group(1)
					element2=re.search("%s/(.*) - (.*)"%self.suitePath,i).group(2)
					contentList.append([element1.strip("\/"),element2.strip("\/")])
		print "failure log has been retrived"
		return sorted(contentList)
		
		
		
	def input_the_value_intoExcel(self):
		for item in self.read_the_summaryValue():
			element1=item[0]
			element2=item[1]
			# print element1,element2,self.rowid,self.maxRow
			if self.maxRow==1:
				self.input_new_row(2,element1,element2)
			else:
				for j in range(self.rowid,self.maxRow+1):
					if self.sheet.cell(column=1,row=j).value==element1:
						if self.sheet.cell(column=2,row=j).value>element2:
							self.sheet.insert_rows(j,1)
							self.input_new_row(j,element1,element2)
							break
						elif self.sheet.cell(column=2,row=j).value==element2:
							self.DupList.append([j,element1,element2])
							self.rowid=j
							self.maxRow=self.sheet.max_row
							break
					elif self.sheet.cell(column=1,row=j).value>element1:
						self.sheet.insert_rows(j,1)
						self.input_new_row(j,element1,element2)
						break
					elif self.sheet.cell(column=1,row=j).value==None:
						self.input_new_row(j,element1,element2)
						break
				else:
					self.input_new_row(j+1,element1,element2)
		self.wb.save(self.excelName)
		
	def printDuplicateList(self):
		#print duplicate log details
		os.system("rm -rf duplicate.txt")
		if self.printDuplicatedFlag and self.DupList!=[]:
			for i in self.DupList:
				print i
		elif self.DupList!=[]:
			print "duplicate details please refer the duplicate.log"
			with open("duplicate.txt","w") as fp:
				fp.write("The below is the duplicate records:\n")
				fp.write("Format is:\n")
				fp.write("[Rowid, SuiteName, CaseName]\n\n")
				for i in self.DupList:
					fp.write(str(i)+'\n')
		else:
			print "there is no duplicate list"
			

		
		
	def input_new_row(self,rowNum,element1,element2):
		self.sheet.cell(column=1,row=rowNum).value=element1
		self.sheet.cell(column=2,row=rowNum).value=element2
		self.sheet.cell(column=self.platformDict[self.platformName[:3]],row=rowNum).value=self.platformName
		for i in range(1,8):
			self.sheet.cell(column=i,row=rowNum).font=Font(name=self.fontName,size=self.fontSize)
			self.sheet.cell(column=i,row=rowNum).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
		self.rowid=rowNum
		self.maxRow=self.sheet.max_row
	
	def parseSummary(self):
		wb=load_workbook(self.excelName)
		sheet=wb["Summary"]
		Dict={}
		for i in range(1,sheet.max_row):
			key=self.sheet.cell(column=1,row=i).value+":"+self.sheet.cell(column=2,row=i).value
			ResolutionValue=self.sheet.cell(column=5,row=i).value
			Dict[key]={"Resolution":ResolutionValue}
			
		return Dict
	
if __name__=="__main__":
	while True:
		print "make sure you have already config the parameter in the property file\n"
		if raw_input("if yes,please press enter key to continue")=="":
			break
	try:
		gs=generateSummary("property.ini")
		gs.input_the_value_intoExcel()
		gs.printDuplicateList()
	except:
		traceback.print_exc()
	if raw_input("please press enter key to terminate")=="":
		sys.exit(0)
