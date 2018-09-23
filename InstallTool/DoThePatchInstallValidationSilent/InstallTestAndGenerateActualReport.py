#encoding=utf-8
import os
import re
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook import Workbook
import collections
import shutil
import sys
import time
import logging
import platform

class silentInstall(object):

	def __init__(self,InstallLogDir,FinalVersion,AffectedFileName,Test='HS',WarnMsgSplit=False,QuickFlag=False):
		self.InstallLogDir=InstallLogDir
		self.InstallFolder=os.getcwd()
		self.InstallIssueList=[]
		self.InstallWarnList=[]
		self.FinalVersion=FinalVersion
		self.AffectedFileName=AffectedFileName
		self.installAnyWhere=['6.2.1','6.2.2']
		self.Test=Test
		self.ModeTestDict={"C":['C-C-C'],"H":['H-H-H'],"S":['S-S-S'],"HS":['H-H-H','H-HS-H','H-HS-HS','S-S-S','S-HS-S','S-HS-HS','C-C-C','C-HS-C','C-HS-HS'],"Quick":['H-H-H','S-S-S','H-H-HS','C-C-C']}
		# self.ModeTestDict={"C":['C-C-C'],"H":['H-H-H'],"S":['S-S-S'],"HS":['H-H-H','H-HS-H','H-HS-HS','S-S-S','S-HS-S','S-HS-HS','C-C-C','C-HS-C','C-HS-HS'],"Quick":['C-C-C']}
		self.BackupFolder='Backup'
		self.FailedFolder='Failed'
		self.NonHSModeVerList=["6.2.0.1","6.2.0.3"]
		self.BaseVer='6.2'
		self.Platform=platform.system()
		self.WarnMsgSplit=WarnMsgSplit
		self.QuickFlag=QuickFlag
		self.StatusIssueList=[]
		self.initFunction()
	def initFunction(self):
		'''initial the logger'''
		if os.path.exists("logbak.txt"):
			os.remove("logbak.txt")
		if os.path.exists("log.txt"):
			os.rename("log.txt","logbak.txt")
		os.system("type nul>log.txt")
		logger = logging.getLogger(__name__)
		logger.setLevel(level = logging.INFO)
		self.logger=logger
		if os.path.exists(os.path.join(self.InstallLogDir,"CISPatchLevel")):
			print "the $HCIROOT %s has CISPatchLevel file, start to uninstall if only the install anywhere version exists"%self.InstallLogDir
			self.logSteps("the $HCIROOT %s has CISPatchLevel file, start to uninstall if only the install anywhere version exists"%self.InstallLogDir)
			address=os.path.join(self.InstallLogDir,"CISPatchLevel")
			fp=open(address,"r")
			content=fp.readlines()
			fp.close()
			for i in content:
				i=i.strip('CIS').strip('\n').strip('.0')
				if i in self.installAnyWhere:
					print "uninstall from "+i+" started"
					self.logSteps("uninstall from "+i+" started")
					if i in self.installAnyWhere:
						os.chdir(os.path.join(self.InstallLogDir,"uninstall_update_%s"%(i+str(".0"))))
						os.system("uninstall_update_%s.exe -i silent" %(i+str(".0")))
						print "uninstall from "+i+" finished"
						self.logSteps("uninstall from "+i+" finished")
				else:
					print "the system has non install anywhere version %s ,please clear the environment, and then reinstall"%i
					self.logSteps("the system has non install anywhere version %s ,please clear the environment, and then reinstall"%i)
					sys.exit(1)

	def logSteps(self,content):
		'''judge whether need to print in the console, and logger the records in the log file'''
		if len(self.logger.handlers)==0:
			handler = logging.FileHandler("log.txt")
			self.logger.addHandler(handler)
		if content.startswith('\n'):
			self.logger.info('\n'+time.strftime("[%m/%d/%Y %H:%M:%S]",time.localtime())+" "+content.strip())
		else:
			self.logger.info(time.strftime("[%m/%d/%Y %H:%M:%S]",time.localtime())+" "+content)
	
	
	
	
	def errorLogJudge(self,fileName,ModeWay,DestiInstallFolder):
		ErrorFlag=True
		WarnFlag=False
		with open(fileName,"r") as fp:
			for i in fp:
				if "successfully" in i or "stallation: Successful" in i:
					ErrorFlag=False
				elif re.match(r"(\d+) Warnings",i):
					if re.match(r"(\d+) Warnings",i).group(1)!='0':
						ErrorFlag=True
						WarnFlag=True
						break
				elif re.match(r"(\d+) NonFatalErrors",i):
					if re.match(r"(\d+) NonFatalErrors",i).group(1)!='0':
						ErrorFlag=True
						break
				elif re.match(r"(\d+) FatalErrors",i):
					if re.match(r"(\d+) FatalErrors",i).group(1)!='0':
						ErrorFlag=True
						break
		if not ErrorFlag:
			print "Move the log file %s is started" %fileName
			self.logSteps("Move the log file %s is started" %fileName)
			shutil.copy(fileName,DestiInstallFolder)
			os.remove(fileName)
			print "Move the log file %s is finished"%fileName
			self.logSteps("Move the log file %s is finished" %fileName)
			return ErrorFlag
		if os.path.exists(ModeWay+"_"+fileName):
			os.remove(ModeWay+"_"+fileName)
		if self.WarnMsgSplit and WarnFlag:
			self.InstallWarnList.append(ModeWay+":"+fileName)
		else:
			self.InstallIssueList.append(ModeWay+":"+fileName)
		self.logSteps(ModeWay+"_"+fileName+" Process Failed")
		os.rename(fileName,ModeWay+"_"+fileName)
		shutil.copy(ModeWay+"_"+fileName,self.FailedFolder)
		os.remove(ModeWay+"_"+fileName)
		return ErrorFlag

	def reNameTheInstallFile(self,orginalLogName,InstallNewLogName,Version,ModeWay):
		try:
			os.rename(orginalLogName,InstallNewLogName)
		except Exception:
			os.chdir(os.path.join(self.InstallFolder,"CIS"+Version.replace(".","")))
			print "it has FatalErrors which results not to continue to install, details please go to the FailLog: %s" %self.FailedFolder
			self.logSteps("it has FatalErrors which results not to continue to install, details please go to the FailLog: %s\n" %self.FailedFolder)
			with open(orginalLogName,"r") as fp:
				for i in fp:
					if "Additional Notes: FATAL ERROR -" in i:
						self.logSteps(i.strip())
			if os.path.exists(ModeWay+"_"+InstallNewLogName):
				os.remove(ModeWay+"_"+InstallNewLogName)
			os.rename(orginalLogName,ModeWay+"_"+InstallNewLogName)
			shutil.copy(ModeWay+"_"+InstallNewLogName,self.FailedFolder)
			sys.exit(1)
			
	def runInstallAnyWhere(self,Version,Mode):
		os.chdir(os.path.join(self.InstallFolder,"CIS"+Version.replace(".","")))
		with open("patch-silent.properties","r") as fpa:
			lines=fpa.readlines()
		with open("patch-silent.properties","w") as fpb:
			for i in lines:
				if re.match(r"\$USER_INSTALL_DIR_SILENT",i):
					fpb.write("$USER_INSTALL_DIR_SILENT$=%s\n" %self.InstallLogDir)
				elif re.match(r"\$CL_INSTALL_TYPE_SILENT",i):
					if Mode.upper()=="H":
						fpb.write("$CL_INSTALL_TYPE_SILENT$=Integration Server\n")
					elif Mode.upper()=="S":
						fpb.write("$CL_INSTALL_TYPE_SILENT$=Security Server\n")
					elif Mode.upper()=="C":
						fpb.write("$CL_INSTALL_TYPE_SILENT$=Client Only\n")
					elif Mode.upper()=="HS":
						fpb.write("$CL_INSTALL_TYPE_SILENT$=Integration Server and Security Server\n")
				else:
					fpb.write(i)
		print "install Version %s is started"%Version
		self.logSteps("install Version %s is started"%Version)
		os.system("CISPatch.bat -i silent -f %s" %os.path.join(os.getcwd(),"patch-silent.properties"))
		print "install Version %s is finished"%Version
		self.logSteps("install Version %s is finished"%Version)

	def runFormerSilentMode(self,Version,Mode):
		os.chdir(os.path.join(self.InstallFolder,"CIS"+Version.replace(".","")))
		print "install Version %s is started"%Version
		self.logSteps("install Version %s is started"%Version)
		if Mode.upper()=="H":
			os.system("install_update -s h")
		elif Mode.upper()=="S":
			os.system("install_update -s s")
		elif Mode.upper()=="C":
			os.system("install_update -s c")
		elif Mode.upper()=="HS":
			os.system("install_update -s hs")
		print "install Version %s is finished"%Version
		self.logSteps("install Version %s is finished"%Version)
	
	
	def startMode(self,StartVersion,StartMode1,StartMode2):
		if StartVersion in self.NonHSModeVerList:
			return StartMode1
		else:
			return StartMode2
			
	def installMethod(self,StartVersion,FinalVersion,StartMode,FinalMode,DestiInstallFolder,ModeWay):
		if StartVersion!=self.BaseVer:
			if StartVersion not in self.installAnyWhere:
				self.runFormerSilentMode(StartVersion,StartMode)
			else:
				self.runInstallAnyWhere(StartVersion,StartMode)
		orginalLogName="integration_update_"+FinalVersion+".0.log"
		InstallNewLogName="install."+FinalVersion+" from "+StartVersion+".log"
		os.chdir(self.InstallLogDir)
		if os.path.exists(orginalLogName):
			os.remove(orginalLogName)
		if os.path.exists(InstallNewLogName):
			os.remove(InstallNewLogName)
		self.runInstallAnyWhere(FinalVersion,FinalMode)
		os.chdir(self.InstallLogDir)
		self.reNameTheInstallFile(orginalLogName,InstallNewLogName,FinalVersion,ModeWay)
		if not self.errorLogJudge(InstallNewLogName,ModeWay,DestiInstallFolder):
			print "uninstall Version %s is started"%FinalVersion
			self.logSteps("uninstall Version %s is started"%FinalVersion)
			os.chdir(os.path.join(self.InstallLogDir,"uninstall_update_%s"%(FinalVersion+str(".0"))))
			os.system("uninstall_update_%s.exe -i silent" %(FinalVersion+str(".0")))
			time.sleep(20)
			print "uninstall Version %s is finished"%FinalVersion
			self.logSteps("uninstall Version %s is finished"%FinalVersion)
			os.chdir(self.InstallLogDir)
			orginalLogName="integration_uninstall_"+FinalVersion+".0.log"
			UninstallNewLogName="uninstall."+FinalVersion+" to "+StartVersion+".log"
			if os.path.exists(UninstallNewLogName):
				os.remove(UninstallNewLogName)
			os.rename(orginalLogName,UninstallNewLogName)
			self.errorLogJudge(UninstallNewLogName,ModeWay,DestiInstallFolder)
			

	#uninstall patch, but it would have one potential issue, if the former patch uninstall issue hasn't found, it would results currpt.
	def uninstall(self,InstalledStartVer,StartMode1,StartMode2):
		if InstalledStartVer[-1:0:-1]!=[]:
			self.logSteps("Start to uninstall the list:%s" %InstalledStartVer[-1:0:-1])
			for i in InstalledStartVer[-1:0:-1]:
				print "uninstall from "+i+" started"
				self.logSteps("uninstall from "+i+" started")
				StartMode=self.startMode(i,StartMode1,StartMode2)
				if i in self.installAnyWhere:
					os.chdir(os.path.join(self.InstallLogDir,"uninstall_update_%s"%(i+str(".0"))))
					os.system("uninstall_update_%s.exe -i silent" %(i+str(".0")))
				else:
					os.chdir(os.path.join(self.InstallFolder,"CIS"+i.replace(".","")))
					if StartMode.upper()=="H":
						os.system("uninstall_update -s h")
					elif StartMode.upper()=="S":
						os.system("uninstall_update -s s")
					elif StartMode.upper()=="C":
						os.system("uninstall_update -s c")
					elif StartMode.upper()=="HS":
						os.system("uninstall_update -s hs")
				print "uninstall from "+i+" finished"
				self.logSteps("uninstall from "+i+" finished")
		
	def judgeModeCheck(self):
		try:
			ModeWay=self.ModeTestDict[self.Test]
		except Exception:
			print "the Test parameter only could be H,S,HS,C"
			sys.exit(0)
		return ModeWay

	def initialBackupFolder(self,FolderName):
		DestiInstallFolder=os.path.join(self.InstallFolder,FolderName)
		if os.path.exists(DestiInstallFolder):
			try:
				shutil.rmtree(os.path.join(self.InstallFolder,self.BackupFolder,FolderName))
			except Exception as e:
				print e.message
				pass
			shutil.move(DestiInstallFolder,os.path.join(self.InstallFolder,self.BackupFolder,FolderName))
		os.mkdir(DestiInstallFolder)
		return DestiInstallFolder

		
	def checkStatus(self,ModeWay,StartVer,FinalMode):
		if self.Platform=="Windows":
			if "H" in FinalMode and os.path.exists(os.path.join(self.InstallLogDir,"server","logs","server.log")):
				os.remove(os.path.join(self.InstallLogDir,"server","logs","server.log"))
			elif "S" in FinalMode and os.path.exists(os.path.join(self.InstallLogDir,"security","logs","security.log")):
				os.remove(os.path.join(self.InstallLogDir,"security","logs","security.log"))
				
			os.system('net start "Infor Cloverleaf(R) Integration Services 62"')
			time.sleep(15)
			if FinalMode=="HS":
				fp=open(os.path.join(self.InstallLogDir,"server","logs","server.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" H")
				fp.close()
				fp=open(os.path.join(self.InstallLogDir,"security","logs","security.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" S")
				fp.close()
			elif FinalMode=="H":
				fp=open(os.path.join(self.InstallLogDir,"server","logs","server.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" H")
				fp.close()
			elif FinalMode=="S":
				fp=open(os.path.join(self.InstallLogDir,"security","logs","security.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" S")
				fp.close()
		else:
			if FinalMode=="HS":
				os.system("hciss -s h")
				os.system("hciss -s s")
				time.sleep(15)
				fp=open(os.path.join(self.InstallLogDir,"server","logs","server.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" H")
				fp.close()
				fp=open(os.path.join(self.InstallLogDir,"security","logs","security.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" S")
				fp.close()
			elif FinalMode=="H":
				os.system("hciss -s h")
				time.sleep(10)
				fp=open(os.path.join(self.InstallLogDir,"server","logs","server.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" H")
				fp.close()
			elif FinalMode=="S":
				os.system("hciss -s s")
				time.sleep(10)
				fp=open(os.path.join(self.InstallLogDir,"security","logs","security.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" S")
				fp.close()
			
	def runInstall(self):
		wb=load_workbook(os.path.join(self.InstallFolder,self.AffectedFileName))
		sheet=wb['Release_Platforms']
		VerList=[]
		DictVer=collections.OrderedDict()
		MaxVerColumn=sheet.max_column
		
		self.FailedFolder=self.initialBackupFolder(self.FailedFolder)
		for i in range(2,MaxVerColumn+1):
			if sheet.cell(column=i,row=1).value:
				VerList.append(str(sheet.cell(column=i,row=1).value))
		MaxRow=sheet.max_row
		for rIndex in range(1,MaxRow+1):
			if re.search(self.Platform,str(sheet.cell(column=1,row=rIndex).value)):
				RowNum=rIndex

		ModeList=["C","H","S"]
		if not self.QuickFlag:
			for index in range(len(VerList)):
				DictVer[VerList[index]]=dict(zip(ModeList,[sheet.cell(column=i,row=RowNum).value for i in range(2+index*3,5+index*3)]))
		else:
			DictVer=collections.OrderedDict([('6.2', {'H': u'\u221a', 'C': u'\u221a', 'S': u'\u221a'}),('6.2.2', {'H': u'\u221a', 'C': u'\u221a', 'S': u'\u221a'})])
		#Run the below mode
		
		ModeWay=self.judgeModeCheck()
		self.logSteps("start to do the installation testing")
		for item1 in ModeWay:
			print "\nThe way "+item1+" installation is started"
			self.logSteps("\nThe way "+item1+" installation is started")
			DestiInstallFolder=self.initialBackupFolder(item1)
			InstalledStartVer=[]
			ModeList=item1.split("-")
			StartMode1=ModeList[0]
			StartMode2=ModeList[1]
			FinalMode=ModeList[2]
			for item2 in DictVer.keys():
				if item2!=self.FinalVersion:
					StartMode=self.startMode(item2,StartMode1,StartMode2)
					# print StartMode,FinalMode
					if(StartMode=="HS" and DictVer[item2]['H'] and DictVer[item2]['S']) or (StartMode!="HS" and DictVer[item2][StartMode]):
						if(FinalMode=="HS" and DictVer[self.FinalVersion]['H'] and DictVer[self.FinalVersion]['S']) or (FinalMode!="HS" and DictVer[self.FinalVersion][FinalMode]):
							print "install from "+item2+" to "+self.FinalVersion+" started"
							self.logSteps("install from "+item2+" to "+self.FinalVersion+" started")
							self.installMethod(item2,self.FinalVersion,StartMode,FinalMode,DestiInstallFolder,item1)
							print "install from "+item2+" to "+self.FinalVersion+" finished"
							self.logSteps("install from "+item2+" to "+self.FinalVersion+" finished")
							InstalledStartVer.append(item2)
							self.checkStatus(item1,item2,FinalMode)
			print "The way "+item1+" installation is finished"
			self.logSteps("The way "+item1+" installation is finished")
			print "The way "+item1+" uninstallation is started"
			self.logSteps("The way "+item1+" uninstallation is started")
			self.uninstall(InstalledStartVer,StartMode1,StartMode2)
			print "The way "+item1+" uninstallation is finished"
			self.logSteps("The way "+item1+" uninstallation is finished")
		if self.InstallIssueList!=[]:
			print "The failed list is %s" %self.InstallIssueList
			self.logSteps("The failed list is %s" %self.InstallIssueList)
			print "Please go to the path %s to find the failed logs"%self.FailedFolder
			self.logSteps("Please go to the path %s to find the failed logs"%self.FailedFolder)
		if self.InstallWarnList!=[]:
			print "The warning log file list is %s" %self.InstallWarnList
			self.logSteps("The warning log file is %s" %self.InstallWarnList)
			print "Please go to the path %s to find the warning logs"%self.FailedFolder
			self.logSteps("Please go to the path %s to find the warning logs"%self.FailedFolder)
		if self.StatusIssueList!=[]:
			print "The status issue list is %s" %self.StatusIssueList
			self.logSteps("The status issue list is %s" %self.StatusIssueList)
		if self.InstallIssueList==[] and self.InstallWarnList==[] and self.StatusIssueList==[]:
			print "The installation process is successfully"
			self.logSteps("The installation process is successfully")

if __name__ == "__main__":
	#InstallLogDir,FinalVersion,AffectedFileName,Test='HS'
	installObj=silentInstall(r"C:\\cloverleaf\\cis6.2\\integrator","6.2.2","CIS.QPD.CIS6.2.X.Affected_File_List.xlsx","HS",True)
	installObj.runInstall()
	
	
