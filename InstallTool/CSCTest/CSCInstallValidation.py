#encoding=utf-8
import os
import re
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook import Workbook
import collections
import shutil
import sys
import time
import logging
import platform
import traceback
from multiprocessing import Pool,Process



class silentInstall(object):

	def __init__(self,debugLevel):
		self.properties=Property("installConfig.ini").parse()
		self.CSC_InstallDir=self.properties['CSC_InstallDir']
		self.CSC_InstallRegDir=self.properties['CSC_InstallRegDir']
		self.CIS_ROOT_Dir=self.properties['CIS_ROOT_Dir']
		self.CSC_Client_Name=self.properties['CSC_Client_Name']
		self.FinalVersion=self.properties['FinalVersion']
		self.AffectedFileName=self.properties['AffectedFileName']
		self.InstallType=self.properties['InstallType']
		self.InstallMode=self.properties['InstallMode']
		self.WarnMsgFilter=self.properties['WarnMsgFilter']
		self.InstallRootFolder=self.properties['InstallRootFolder']
		self.Build=self.properties['Build']
		self.debugLevel=debugLevel
		self.InstallToolFolder=os.getcwd()
		self.InstallIssueList=[]
		self.InstallWarnList=[]
		self.BaseVer=None
		self.installAnyWhere=['6.3']
		self.BackupFolder='Backup'
		self.FailedFolder='Failed'
		self.SuccessfulFolder='Succeeded'
		self.Platform=platform.system()
		if self.Platform=="Windows":
			import win32api
			import win32con
		self.StatusIssueList=[]
		self.initFunction()
		self.rt=runtime(self.InstallToolFolder)
		
	def initFunction(self):
		'''initial the logger'''
		if os.path.exists("logbak.txt"):
			os.remove("logbak.txt")
		if os.path.exists("log.txt"):
			os.rename("log.txt","logbak.txt")
		if self.Platform=="Windows":
			os.system("type nul>log.txt")
		else:
			os.system("touch log.txt")
		logger = logging.getLogger(__name__)
		logger.setLevel(level = logging.INFO)
		self.logger=logger
		if self.Platform=="Windows":
			self.KillPid("java.exe")
			self.KillPid("tomcat9.exe")
		if self.debugLevel!="debug":
			self.uninstall_exist_csc_server()
			self.uninstall_exist_csc_client()
		
	
	
	def uninstall_exist_csc_server(self):
		if os.path.exists(os.path.join(self.CSC_InstallDir,"server","version")):
			print "the $CSC_SERVER_HOME %s has version file, start to uninstall it"%os.path.join(self.CSC_InstallDir,"server")
			self.logSteps("the $CSC_SERVER_HOME %s has version file, start to uninstall it"%os.path.join(self.CSC_InstallDir,"server"))
			address=os.path.join(self.CSC_InstallDir,"server","version")
			fp=open(address,"r")
			content=fp.readlines()
			fp.close()
			if len(str(content).strip().split())==1:
				os.system("%s -i silent"%os.path.join(self.CSC_InstallDir,"server","uninstall_csc_server","uninstall_csc_server"))
			else:
				for i in content:
					i=i.strip('\n')
					print "uninstall from "+i+" started"
					self.logSteps("uninstall from "+i+" started")
					if i in self.installAnyWhere:
						os.chdir(os.path.join(self.CSC_InstallDir,"server","uninstall_update_%s"%(i)))
						os.system("uninstall_update_%s.exe -i silent" %(i))
					else:
						os.chdir(os.path.join(self.InstallRootFolder,"CSC"+i.replace(".","")))
						os.system("uninstall_server_update -s")
					print "uninstall from "+i+" finished"
					self.logSteps("uninstall from "+i+" finished")
			os.system("rm -rf %s"%os.path.join(self.CSC_InstallDir,"server"))
			print "finish the uninstall for the exist CSC Server"
			self.logSteps("finish the uninstall for exist the CSC Server")
	
	def uninstall_exist_csc_client(self):
		if os.path.exists(os.path.join(self.CSC_InstallDir,"client_%s"%self.CSC_Client_Name,"version")):
			print "the $CSC_CLIENT_HOME %s has version file, start to uninstall it"%os.path.join(self.CSC_InstallDir,"client-%s"%self.CSC_Client_Name)
			self.logSteps("the $CSC_CLIENT_HOME %s has version file, start to uninstall it"%os.path.join(self.CSC_InstallDir,"client-%s"%self.CSC_Client_Name))
			address=os.path.join(self.CSC_InstallDir,"client_%s"%self.CSC_Client_Name,"version")
			fp=open(address,"r")
			content=fp.readlines()
			fp.close()
			if len(str(content).strip().split())==1:
				os.system("%s -i silent"%os.path.join(self.CSC_InstallDir,"client_%s"%self.CSC_Client_Name,"uninstall_csc_client","uninstall_csc_client"))
			else:
				for i in content:
					i=i.strip('\n')
					print "uninstall from "+i+" started"
					self.logSteps("uninstall from "+i+" started")
					os.chdir(os.path.join(self.CSC_InstallDir,"client_%s"%self.CSC_Client_Name,"uninstall_update_%s"%(i)))
					os.system("uninstall_update_%s.exe -i silent" %(i))
					print "uninstall from "+i+" finished"
					self.logSteps("uninstall from "+i+" finished")
			os.system("rm -rf %s"%os.path.join(self.CSC_InstallDir,"client_%s"%self.CSC_Client_Name))
			print "finish the uninstall for the exist CSC Client"
			self.logSteps("finish the uninstall for the exist CSC Client")
	
	def logSteps(self,content):
		'''judge whether need to print in the console, and logger the records in the log file'''
		if len(self.logger.handlers)==0:
			handler = logging.FileHandler("log.txt")
			self.logger.addHandler(handler)
		if content.startswith('\n'):
			self.logger.info('\n'+time.strftime("[%m/%d/%Y %H:%M:%S]",time.localtime())+" "+content.strip())
		else:
			self.logger.info(time.strftime("[%m/%d/%Y %H:%M:%S]",time.localtime())+" "+content)
	
	
	
	
	def errorLogJudge(self,fileName,DestiInstallFolder):
		ErrorFlag=False
		WarnCount=0
		with open(fileName,"r") as fp:
			for i in fp:
				if "stallation: Successful." in i:
					break
				elif re.match(r"(\d+) Warnings",i):
					if re.match(r"(\d+) Warnings",i).group(1)!='0':
						WarnCount=int(re.match(r"(\d+) Warnings",i).group(1))
				elif re.match(r"(\d+) NonFatalErrors",i):
					if re.match(r"(\d+) NonFatalErrors",i).group(1)!='0':
						ErrorFlag=True
				elif re.match(r"(\d+) FatalErrors",i):
					if re.match(r"(\d+) FatalErrors",i).group(1)!='0':
						ErrorFlag=True
		
		#move the successfull log to the succeed log folder if there is no error or warning
		WarnCount=self.removeNormalWarn(WarnCount,fileName)
		if self.WarnMsgFilter:
			if not ErrorFlag and WarnCount==0:
				self.moveSuccessfulFile(fileName,DestiInstallFolder)
				return ErrorFlag
		else:
			if not ErrorFlag:
				self.moveSuccessfulFile(fileName,DestiInstallFolder)
				return ErrorFlag
		#move the issue to the failed folder
		if ErrorFlag:
			self.InstallIssueList.append(fileName)
		if self.WarnMsgFilter and WarnCount!=0:
			self.InstallWarnList.append(fileName)
		self.logSteps(fileName+" Process Failed")
		shutil.copy(fileName,self.FailedFolder)
		os.remove(fileName)
		return ErrorFlag

	def removeNormalWarn(self,WarnCount,fileName):
		if WarnCount!=0:
			fp=open(fileName,"r")
			content=fp.read()
			if "csc_client_install.log\n                          Status: WARNING" in content:
				WarnCount-=1
			if "csc_server_install.log\n                          Status: WARNING" in content:
				WarnCount-=1
			if "SetupServerCert.bat\n                          Status: WARNING" in content:
				WarnCount-=1
				
		return WarnCount
		
		
	def moveSuccessfulFile(self,fileName,DestiInstallFolder):
		print "Move the log file %s is started" %fileName
		self.logSteps("Move the log file %s is started" %fileName)
		shutil.copy(fileName,DestiInstallFolder)
		os.remove(fileName)
		print "Move the log file %s is finished"%fileName
		self.logSteps("Move the log file %s is finished" %fileName)
		
		
	def switch_csc_root(self,Type):
		if Type.lower()=="client":
			os.chdir(os.path.join(self.CSC_InstallDir,Type.lower()+"_"+self.CSC_Client_Name))
		else:
			os.chdir(os.path.join(self.CSC_InstallDir,Type.lower()))
	
	def reNameTheInstallFile(self,orginalLogName,InstallNewLogName,Version,Type,Mode):
		self.switch_csc_root(Type)
		if os.path.exists(InstallNewLogName):
			os.remove(InstallNewLogName)
		if os.path.exists(orginalLogName):
			os.rename(orginalLogName,InstallNewLogName)
		else:
			os.chdir(os.path.join(self.InstallRootFolder,"CSC"+Version.replace(".",""),self.Build,Type.lower(),Mode))
			print "it has FatalErrors which results not to continue to install, details please go to the FailLog: %s" %self.FailedFolder
			self.logSteps("it has FatalErrors which results not to continue to install, details please go to the FailLog: %s\n" %self.FailedFolder)
			with open(orginalLogName,"r") as fp:
				for i in fp:
					if "Additional Notes: FATAL ERROR -" in i:
						self.logSteps(i.strip())
			if os.path.exists(InstallNewLogName):
				os.remove(InstallNewLogName)
			os.rename(orginalLogName,InstallNewLogName)
			shutil.copy(InstallNewLogName,self.FailedFolder)
			sys.exit(1)
			
	def runInstallAnyWherePatch(self,Version,Mode):
		os.chdir(os.path.join(self.InstallRootFolder,"CSC"+Version.replace(".",""),self.Build,"server","Patch"))
		with open("serversilent.properties","r") as fpa:
			lines=fpa.readlines()
		with open("serversilent.properties","w") as fpb:
			for i in lines:
				if re.match(r"\$USER_INSTALL_DIR_SILENT",i):
					fpb.write("$USER_INSTALL_DIR_SILENT$=%s\n" %self.InstallLogDir)
				elif re.match(r"\$CL_INSTALL_TYPE_SILENT",i):
					fpb.write()
				else:
					fpb.write(i)
		print "install Version %s is started"%Version
		self.logSteps("install Version %s is started"%Version)
		os.system("CISPatch.bat -i silent -f %s" %os.path.join(os.getcwd(),"patch-silent.properties"))
		print "install Version %s is finished"%Version
		self.logSteps("install Version %s is finished"%Version)

	def runFormerSilentMode(self,Version,Mode):
		os.chdir(os.path.join(self.InstallRootFolder,"CSC"+Version.replace(".",""),self.Build,"server","Patch"))
		print "install Version %s is started"%Version
		self.logSteps("install Version %s is started"%Version)
		os.system("install_server_update -s")
		print "install Version %s is finished"%Version
		self.logSteps("install Version %s is finished"%Version)
	
	def csc_full_install(self,Type,Version,DestiInstallFolder):
		os.chdir(os.path.join(self.InstallRootFolder,"CSC"+Version.replace(".",""),self.Build,Type.lower(),"Full"))
		with open("%ssilent.properties"%Type.lower(),"r") as fpa:
			lines=fpa.readlines()
		with open("%ssilent.properties"%Type.lower(),"w") as fpb:
			for i in lines:
				if re.match(r"  USER_INSTALL_DIR",i):
					fpb.write("  USER_INSTALL_DIR=%s\n" %self.CSC_InstallDir)
				elif re.match(r"  CLIENT_NAME",i):
					fpb.write("  CLIENT_NAME=%s\n" %self.CSC_Client_Name)
				elif re.match(r"  CL_SITE_PATH",i):
					fpb.write("  CL_SITE_PATH=%s\n" %self.CIS_ROOT_Dir)
				elif re.match(r"  USER_REQUESTED_RESTART",i):
					fpb.write('  USER_REQUESTED_RESTART=\"NO\"\n')
				else:
					fpb.write(i)
		print "install Version %s %s is started"%(Version,Type)
		self.logSteps("install Version %s is started"%Version)
		orginalLogName="csc_%s_install.log"%Type.lower()
		InstallNewLogName=Type+"_"+"Full"+"_"+orginalLogName
		if os.path.exists(orginalLogName):
			os.remove(orginalLogName)
		if self.Platform=="Windows":
			os.system("CSC.bat -i silent -f %s" %os.path.join(os.getcwd(),"%ssilent.properties"%Type.lower()))
		else:
			if Type=="client":
				application="CSCClient.bin"
			elif Type=="server":
				application="csc_server_install.bin"
			os.system("%s -i silent -f %s" %(os.path.join(os.getcwd(),application),os.path.join(os.getcwd(),"%ssilent.properties"%Type.lower())))
		print "install Version %s %s is finished"%(Version,Type)
		self.logSteps("install Version %s is finished"%Version)
		self.reNameTheInstallFile(orginalLogName,InstallNewLogName,Version,Type.lower(),"Full")
		self.errorLogJudge(InstallNewLogName,DestiInstallFolder)
		
	def csc_full_uninstall(self,Type,Version,DestiInstallFolder):
		self.switch_csc_root(Type)
		os.chdir(os.path.join(os.getcwd(),"uninstall_csc_%s"%Type.lower()))
		print "uninstall %s CSC Version %s is started"%(Type,Version)
		self.logSteps("uninstall %s CSC Version %s is started"%(Type,Version))
		orginalLogName="csc_%s_uninstall.log"%Type.lower()
		InstallNewLogName=Type+"_"+"Full"+"_"+orginalLogName
		if os.path.exists(orginalLogName):
			os.remove(orginalLogName)
		os.system("uninstall_csc_%s -i silent"%Type.lower())
		print "uninstall Version %s is finished"%Version
		self.logSteps("uninstall Version %s is finished"%Version)
		self.reNameTheInstallFile(orginalLogName,InstallNewLogName,Version,Type.lower(),"Full")
		self.errorLogJudge(InstallNewLogName,DestiInstallFolder)
		
		
	def copyBatFile(self,Type,Version,Mode):
		os.chdir(self.InstallToolFolder)
		if Mode=="Full":
			File="CSCFull.bat"
		elif Mode=="Patch":
			File="CSCPatch.bat"
		TargetPath=os.path.join(self.InstallRootFolder,"CSC"+Version.replace(".",""),self.Build,Type.lower(),Mode)
		os.system('cp -r -f %s %s'%(File,TargetPath))
		if os.path.exists(os.path.join(TargetPath,"CSC.bat")):
			os.remove(os.path.join(TargetPath,"CSC.bat"))
		os.rename(os.path.join(TargetPath,File),os.path.join(TargetPath,"CSC.bat"))
		with open(os.path.join(TargetPath,"CSC.bat"),"r") as fpa:
			lines=fpa.readlines()
		if Type.lower()=="server":
			InstallFile="csc_server_install.exe"
		elif Type.lower()=="client":
			InstallFile="CSCClient.exe"
		with open(os.path.join(TargetPath,"CSC.bat"),"w") as fpb:
			for i in lines:
				if re.match(r"start csc",i):
					fpb.write("start %s %%*"%InstallFile)
				elif re.match(r"csc",i) or re.match(r"CSC",i):
					fpb.write("%s %%*"%InstallFile)
				else:
					fpb.write(i)
		
	def csc_patch_install(self,StartVersion,FinalVersion,DestiInstallFolder):
		if StartVersion==self.BaseVer:
			self.csc_full_install("server",self.BaseVer)
		else:
			if StartVersion not in self.installAnyWhere:
				self.runFormerSilentMode(StartVersion)
			else:
				self.runInstallAnyWherePatch(StartVersion)
		orginalLogName="csc_server_update_"+FinalVersion+".log"
		InstallNewLogName="Server_Patch_"+"install."+FinalVersion+" from "+StartVersion+".log"
		os.chdir(os.path.join(self.CSC_InstallDir,"server"))
		if os.path.exists(orginalLogName):
			os.remove(orginalLogName)
		if os.path.exists(InstallNewLogName):
			os.remove(InstallNewLogName)
		self.runInstallAnyWherePatch(FinalVersion,FinalMode)
		os.chdir(os.path.join(self.CSC_InstallDir,"server"))
		self.reNameTheInstallFile(orginalLogName,InstallNewLogName,FinalVersion,"Patch")
		if not self.errorLogJudge(InstallNewLogName,DestiInstallFolder):
			print "uninstall Version %s is started"%FinalVersion
			self.logSteps("uninstall Version %s is started"%FinalVersion)
			os.chdir(os.path.join(self.CSC_InstallDir,"server","uninstall_server_update_%s"%(FinalVersion)))
			os.system("csc_sever_update_%s.exe -i silent" %(FinalVersion))
			# time.sleep(3)
			print "uninstall Version %s is finished"%FinalVersion
			self.logSteps("uninstall Version %s is finished"%FinalVersion)
			os.chdir(self.InstallLogDir)
			orginalLogName="csc_server_uninstall_"+FinalVersion+".log"
			UninstallNewLogName="Server_Patch_"+"uninstall."+FinalVersion+" to "+StartVersion+".log"
			if os.path.exists(UninstallNewLogName):
				os.remove(UninstallNewLogName)
			os.rename(orginalLogName,UninstallNewLogName)
			self.errorLogJudge(UninstallNewLogName,DestiInstallFolder)
		
		
		
		
	#uninstall patch, but it would have one potential issue, if the former patch uninstall issue hasn't found, it would results currpt.
	def uninstall(self,InstalledStartVer):
		if InstalledStartVer[-1:0:-1]!=[]:
			self.logSteps("Start to uninstall the list:%s" %InstalledStartVer[-1:0:-1])
			for i in InstalledStartVer[-1:0:-1]:
				print "uninstall from "+i+" started"
				self.logSteps("uninstall from "+i+" started")
				if i in self.installAnyWhere:
					os.chdir(os.path.join(self.InstallLogDir,"uninstall_update_%s"%(i)))
					os.system("uninstall_update_%s.exe -i silent" %(i))
				else:
					os.chdir(os.path.join(self.InstallRootFolder,"CSC"+i.replace(".","")))
					os.system("uninstall_server_update -s")
				print "uninstall from "+i+" finished"
				self.logSteps("uninstall from "+i+" finished")
		

	def initialBackupFolder(self,FolderName):
		DestiInstallFolder=os.path.join(self.InstallToolFolder,FolderName)
		if os.path.exists(DestiInstallFolder):
			try:
				shutil.rmtree(os.path.join(self.InstallToolFolder,self.BackupFolder,FolderName))
			except Exception as e:
				print e.message
				pass
			shutil.move(DestiInstallFolder,os.path.join(self.InstallToolFolder,self.BackupFolder,FolderName))
		os.mkdir(DestiInstallFolder)
		return DestiInstallFolder

		
	def checkStatus(self,ModeWay,StartVer,FinalMode):
		if self.Platform=="Windows":
			if "H" in FinalMode and os.path.exists(os.path.join(self.InstallLogDir,"server","logs","server.log")):
				os.remove(os.path.join(self.InstallLogDir,"server","logs","server.log"))
			elif "S" in FinalMode and os.path.exists(os.path.join(self.InstallLogDir,"security","logs","security.log")):
				os.remove(os.path.join(self.InstallLogDir,"security","logs","security.log"))
				
			os.system('net start "Infor Cloverleaf(R) Integration Services 62"')
			time.sleep(15)
			if FinalMode=="HS":
				fp=open(os.path.join(self.InstallLogDir,"server","logs","server.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" H")
				fp.close()
				fp=open(os.path.join(self.InstallLogDir,"security","logs","security.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" S")
				fp.close()
			elif FinalMode=="H":
				fp=open(os.path.join(self.InstallLogDir,"server","logs","server.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" H")
				fp.close()
			elif FinalMode=="S":
				fp=open(os.path.join(self.InstallLogDir,"security","logs","security.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" S")
				fp.close()
		else:
			if FinalMode=="HS":
				os.system("hciss -s h")
				os.system("hciss -s s")
				time.sleep(15)
				fp=open(os.path.join(self.InstallLogDir,"server","logs","server.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" H")
				fp.close()
				fp=open(os.path.join(self.InstallLogDir,"security","logs","security.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" S")
				fp.close()
			elif FinalMode=="H":
				os.system("hciss -s h")
				time.sleep(10)
				fp=open(os.path.join(self.InstallLogDir,"server","logs","server.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" H")
				fp.close()
			elif FinalMode=="S":
				os.system("hciss -s s")
				time.sleep(10)
				fp=open(os.path.join(self.InstallLogDir,"security","logs","security.log"),"r")
				if "Started Successfully!" not in fp.read():
					self.StatusIssueList.append(ModeWay+":"+"install from "+StartVer+" to "+self.FinalVersion+" S")
				fp.close()
			
			
			
	def runInstall(self):
		wb=load_workbook(os.path.join(self.InstallToolFolder,self.AffectedFileName))
		sheet=wb['Release_Platforms']
		VerList=[]
		MaxVerColumn=sheet.max_column
		self.FailedFolder=self.initialBackupFolder(self.FailedFolder)
		DestiInstallFolder=self.initialBackupFolder(self.SuccessfulFolder)
		#this is for version loop installation
		for i in range(2,MaxVerColumn+1):
			if sheet.cell(column=i,row=1).value:
				VerList.append(str(sheet.cell(column=i,row=1).value))
		print VerList
		self.BaseVer=VerList[0]
		#Run the below mode
		if self.InstallType.upper()=="ALL":
			if self.Platform=="Windows":
				self.copyBatFile("server",self.FinalVersion,"Full")
				self.copyBatFile("client",self.FinalVersion,"Full")
			if self.InstallMode.upper()=="ALL":
				self.csc_full_install("server",self.FinalVersion,DestiInstallFolder)
				self.csc_full_install("client",self.FinalVersion,DestiInstallFolder)
				self.csc_full_uninstall("server",self.FinalVersion,DestiInstallFolder)
				self.csc_full_uninstall("client",self.FinalVersion,DestiInstallFolder)
				for i in VerList:
					self.csc_patch_install()
				for i in VerList[::-1]:
					self.csc_patch_uninstall()
				self.csc_full_uninstall("server",VerList[0],DestiInstallFolder)
			elif self.InstallMode.upper()=="FULL":
				self.csc_full_install("server",self.FinalVersion,DestiInstallFolder)
				self.csc_full_install("client",self.FinalVersion,DestiInstallFolder)
				self.initialRatsEnvironment()
				if self.Platform=="Windows":
					self.serverWinRegExists()
					self.clientWinRegExists()
					self.rt.running()
					self.rt.checkSanityIssue()
					self.csc_full_uninstall("server",self.FinalVersion,DestiInstallFolder)
					self.csc_full_uninstall("client",self.FinalVersion,DestiInstallFolder)
					self.serverWinRegNonExists()
					self.clientWinRegNonExists()
				else:
					self.linuxServiceExists()
			elif self.InstallMode.upper()=="PATCH":
				for i in VerList:
					self.csc_patch_install()
				for i in VerList[::-1]:
					self.csc_patch_uninstall()
				self.csc_full_uninstall("server",VerList[0],DestiInstallFolder)
					
		elif self.InstallType.upper()=="SERVER":
			self.copyBatFile("server",self.FinalVersion,"Full")
			if self.InstallMode.upper()=="ALL":
				self.csc_full_install("server",self.FinalVersion,DestiInstallFolder)
				self.csc_full_uninstall("server",self.FinalVersion,DestiInstallFolder)
				for i in VerList:
					self.csc_patch_install()
				for i in VerList[::-1]:
					self.csc_patch_uninstall()
				self.csc_full_uninstall("server",VerList[0],DestiInstallFolder)
			elif self.InstallMode.upper()=="FULL":
				self.csc_full_install("server",self.FinalVersion,DestiInstallFolder)
				self.csc_full_uninstall("server",self.FinalVersion,DestiInstallFolder)
			elif self.InstallMode.upper()=="PATCH":
				for i in VerList:
					self.csc_patch_install()
				for i in VerList[::-1]:
					self.csc_patch_uninstall()
				self.csc_full_uninstall("server",VerList[0])
		elif self.InstallType.upper()=="CLIENT":
			self.copyBatFile("client",self.FinalVersion,"Full")
			if self.InstallMode.upper()=="ALL" or self.InstallMode.upper()=="FULL":
				self.csc_full_install("client",self.FinalVersion,DestiInstallFolder)
				self.csc_full_uninstall("client",self.FinalVersion,DestiInstallFolder)
				
		#The final status:
		if self.InstallIssueList!=[]:
			print "The failed list is %s" %self.InstallIssueList
			self.logSteps("The failed list is %s" %self.InstallIssueList)
			print "Please go to the path %s to find the failed logs"%self.FailedFolder
			self.logSteps("Please go to the path %s to find the failed logs"%self.FailedFolder)
		if self.InstallWarnList!=[]:
			print "The warning log file list is %s" %self.InstallWarnList
			self.logSteps("The warning log file is %s" %self.InstallWarnList)
			print "Please go to the path %s to find the warning logs"%self.FailedFolder
			self.logSteps("Please go to the path %s to find the warning logs"%self.FailedFolder)
		if self.StatusIssueList!=[]:
			print "The status issue list is %s" %self.StatusIssueList
			self.logSteps("The status issue list is %s" %self.StatusIssueList)
		if self.InstallIssueList==[] and self.InstallWarnList==[] and self.StatusIssueList==[]:
			print "The installation process is successfully"
			self.logSteps("The installation process is successfully")
	
	def initialRatsEnvironment(self):
		clSampleSitePath=os.path.join(self.CSC_InstallDir,"server","sample_sites","64bit","csc_cloverleaf")
		srSampleSitePath=os.path.join(self.CSC_InstallDir,"server","sample_sites","64bit","csc_server")
		os.system("cp -r -f %s %s"%(clSampleSitePath,os.path.join(self.CSC_InstallDir,"server","sample_sites")))
		os.system("cp -r -f %s %s"%(srSampleSitePath,os.path.join(self.CSC_InstallDir,"server","sample_sites")))
		os.system("cp -r -f %s %s"%(os.path.join(self.InstallToolFolder,"license.dat"),os.path.join(self.CSC_InstallDir,"server","vers")))
		with open(os.path.join(self.CSC_InstallDir,"server","tomcat","engine","conf","catalina.policy"),"r") as fpa:
			content=fpa.readlines()
		with open(os.path.join(self.CSC_InstallDir,"server","tomcat","engine","conf","catalina.policy"),"w") as fpb:
			for i in content:
				if i.startswith(r'grant codeBase "file:${catalina.base}/regapps/regwebsvc/-" {') or i.startswith(r'grant codeBase "file:${catalina.base}/webapps/cscwebsvc/-" {'):
					fpb.write(i+'    permission java.io.FilePermission "%s/-", "read,write,delete";\n'%self.CIS_ROOT_Dir.replace("\\\\","/"))
				else:
					fpb.write(i)
	
	def KillPid(self,app,details=None):
		os.chdir(self.InstallToolFolder)
		if self.Platform=="Windows":
			os.system("tasklist>taskList.txt")
			List=[]
			with open("taskList.txt","r") as fp:
				for line in fp:
					if app in line and re.findall(r"\s(\d+)\s",line,re.M)[0] not in List:
						List.append(re.findall(r"\s(\d+)\s",line,re.M)[0])
			for item in List:
				os.system("taskkill/F /pid %s" %item)
		else:
			os.system("ps -ef|grep %s>taskList.txt"%app)
			with open("taskList.txt","r") as fp:
				for line in fp:
					if details==None:
						os.system("kill -9 %s"%line.split()[1])
					else:
						if details in line:
							os.system("kill -9 %s"%line.split()[1])
		
	def compareReg(self,path,key,expected):
		result=os.system('REG QUERY "%s" -v %s>reg.txt'%(path,key))
		with open("reg.txt","r") as fp:
			content=fp.readlines()
		if result==0:
			value=re.search(r"REG_SZ(.*)","".join(content)).group(1).strip()
			if value==expected:
				return "True"
			else:
				return value
		else:
			return content
		
	def linuxServiceExists(self):
		fp=open(os.path.join(self.CSC_InstallDir,"client_%s"%self.CSC_Client_Name,"dispatcher","dispatcher.log"),"r")
		if "Running dispatcher command server on port" not in fp.read():
			self.printLog("the client dispatcher hasn't been started")
		else:
			self.printLog("the client dispatcher has already been started")
		fp.close()
		os.system("su hci -l -c %s"%os.path.join(self.CSC_InstallDir,"server","lws","service","start_csc_server.sh"))
		time.sleep(25)
		os.system("%s>status.txt"%os.path.join(self.CSC_InstallDir,"server","bin","cscserverstatus"))
		fp=open("status.txt","r")
		if "CSC Server service is running" not in fp.read():
			self.printLog("the server service hasn't been started")
		else:
			self.printLog("the server service has already been started")
		fp.close()
	
	
	def serverWinRegExists(self):
		#comare the control panel program name
		result=self.compareReg("HKEY_LOCAL_MACHINE\SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall\Infor Cloverleaf(R) Secure Courier %s Server"%self.FinalVersion.strip(".0"),"DisplayName","Infor Cloverleaf(R) Secure Courier %s Server"%self.FinalVersion.strip(".0"))
		if result!="True":
			self.printLog("server control panel comparison failed\n,the actual value is %s,the expected value is %s"%(result,"Infor Cloverleaf(R) Secure Courier %s Server"%self.FinalVersion.strip(".0")))
		else:
			self.printLog("the server control panel exists check successfully")
		#comare the install place
		result=self.compareReg("HKEY_LOCAL_MACHINE\SOFTWARE\Infor\Infor Cloverleaf(R) Secure Courier %s Server"%self.FinalVersion.strip(".0"),"InstallLocation",os.path.join(self.CSC_InstallRegDir,"server"))
		if result!="True":
			self.printLog("server install place regdit comparison failed\n,the actual value is %s,the expected value is %s"%(result,os.path.join(self.CSC_InstallRegDir,"server")))
		else:
			self.printLog("the server install regdit exists check successfully")
		#comare the service name
		result=self.compareReg("HKEY_LOCAL_MACHINE\SYSTEM\ControlSet001\Services\CSCServer%s"%self.FinalVersion.strip(".0").replace(".",""),"DisplayName","Infor Cloverleaf(R) Secure Courier %s Server"%self.FinalVersion.strip(".0"))
		if result!="True":
			self.printLog("server service name comparison failed\n,the actual value is %s,the expected value is %s"%(result,"Infor Cloverleaf(R) Secure Courier %s Server"%self.FinalVersion.strip(".0")))
		else:
			self.printLog("the server service exists check successfully")
		#start menu compare
		if not self.checkStartMenu(os.path.join(r"C:\ProgramData\Microsoft\Windows\Start Menu\Programs","Infor Cloverleaf Secure Courier %s Server"%self.FinalVersion.strip(".0"))):
			self.printLog("server start menu Infor Cloverleaf Secure Courier %s Server couldn't be found"%self.FinalVersion.strip(".0"))
		else:
			self.printLog("the server startMenu exists check successfully")
		
		
	def clientWinRegExists(self):
		#comare the control panel program name
		result=self.compareReg("HKEY_LOCAL_MACHINE\SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall\Infor Cloverleaf(R) Secure Courier %s Client"%self.FinalVersion.strip(".0"),"DisplayName","Infor Cloverleaf(R) Secure Courier %s Client"%self.FinalVersion.strip(".0"))
		if result!="True":
			self.printLog("client control panel comparison failed\n,the actual value is %s,the expected value is %s"%(result,"Infor Cloverleaf(R) Secure Courier %s Client"%self.FinalVersion.strip(".0")))
		else:
			self.printLog("the client control panel exists check successfully")
		#comare the install place
		result=self.compareReg("HKEY_LOCAL_MACHINE\SOFTWARE\Infor\Infor Cloverleaf(R) Secure Courier %s Client %s"%(self.FinalVersion.strip(".0"),self.CSC_Client_Name),"InstallLocation",os.path.join(self.CSC_InstallRegDir,"client_%s"%self.CSC_Client_Name))
		if result!="True":
			self.printLog("client install place regdit comparison failed\n,the actual value is %s,the expected value is %s"%(result,os.path.join(self.CSC_InstallRegDir,"client_%s"%self.CSC_Client_Name)))
		else:
			self.printLog("the client install regdit exists check successfully")
		#comare the service name
		result=self.compareReg("HKEY_LOCAL_MACHINE\SYSTEM\ControlSet001\Services\CSCClient%s%s"%(self.FinalVersion.strip(".0").replace(".",""),self.CSC_Client_Name),"DisplayName","Infor Cloverleaf(R) Secure Courier %s Client %s"%(self.FinalVersion.strip(".0"),self.CSC_Client_Name))
		if result!="True":
			self.printLog("client service name comparison failed\n,the actual value is %s,the expected value is %s"%(result,"Infor Cloverleaf(R) Secure Courier %s Client %s"%(self.FinalVersion.strip(".0"),self.CSC_Client_Name)))
		else:
			self.printLog("the client service exists check successfully")
		
		#start menu compare
		if not self.checkStartMenu(os.path.join(r"C:\ProgramData\Microsoft\Windows\Start Menu\Programs","Infor Cloverleaf Secure Courier %s Client %s"%(self.FinalVersion.strip(".0"),self.CSC_Client_Name))):
			self.printLog("client start menu Infor Cloverleaf Secure Courier %s Client %s couldn't be found"%(self.FinalVersion.strip(".0"),self.CSC_Client_Name))
		else:
			self.printLog("the client startMenu exists check successfully")
		
	def logFileContent(self,fileName):
		with open(fileName,"r") as fp:
			for line in fp:
				self.printLog(line)
		
		
	def serverWinRegNonExists(self):
		#check the control panel program name is removed
		result=os.system('REG QUERY "HKEY_LOCAL_MACHINE\SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall\Infor Cloverleaf(R) Secure Courier %s Server">error.txt'%self.FinalVersion.strip(".0"))
		if result==0:
			self.printLog("server control panel has't been cleared successfully\n,the value Infor Cloverleaf(R) Secure Courier %s Server still exists"%"Infor Cloverleaf(R) Secure Courier %s Client"%self.FinalVersion.strip(".0"))
			self.logFileContent("error.txt")
		else:
			self.printLog("server control panel has been removed successfully")
		#check the install place is removed
		result=os.system('REG QUERY "HKEY_LOCAL_MACHINE\SOFTWARE\Infor\Infor Cloverleaf(R) Secure Courier %s Server">error.txt'%self.FinalVersion.strip(".0"))
		if result==0:
			self.printLog("server install place regdit hasn't been cleared successfully\n,the value Infor Cloverleaf(R) Secure Courier %s Server still exists"%self.FinalVersion.strip(".0"))
			self.logFileContent("error.txt")
		else:
			self.printLog("server install place regdit has been removed successfully")
		#check the service name is removed
		result=os.system('REG QUERY "HKEY_LOCAL_MACHINE\SYSTEM\ControlSet001\Services\CSCServer%s">error.txt'%self.FinalVersion.strip(".0").replace(".",""))
		if result==0:
			self.printLog("server service name hasn't been cleared successfully\n,the value CSCServer%s still exists"%self.FinalVersion.strip(".0").replace(".",""))
			self.logFileContent("error.txt")
		else:
			self.printLog("server service name has been removed successfully")
		#start menu compare
		if self.checkStartMenu(os.path.join(r"C:\ProgramData\Microsoft\Windows\Start Menu\Programs","Infor Cloverleaf Secure Courier %s Server"%self.FinalVersion.strip(".0"))):
			self.printLog("server start menu Infor Cloverleaf Secure Courier %s Server still could be found"%self.FinalVersion.strip(".0"))
		else:
			self.printLog("server start menu has been removed successfully")
			
	def clientWinRegNonExists(self):
		#check the control panel program name is removed
		result=os.system('REG QUERY "HKEY_LOCAL_MACHINE\SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall\Infor Cloverleaf(R) Secure Courier %s Client">error.txt'%self.FinalVersion.strip(".0"))
		if result==0:
			self.printLog("client control panel has't been cleared successfully\n,the value Infor Cloverleaf(R) Secure Courier %s Client still exists"%"Infor Cloverleaf(R) Secure Courier %s Client"%self.FinalVersion.strip(".0"))
			self.logFileContent("error.txt")
		else:
			self.printLog("client control panel has been cleared successfully")
		#check the install place is removed
		result=os.system('REG QUERY "HKEY_LOCAL_MACHINE\SOFTWARE\Infor\Infor Cloverleaf(R) Secure Courier %s Client %s">error.txt'%(self.FinalVersion.strip(".0"),self.CSC_Client_Name))
		if result==0:
			self.printLog("client install place regdit hasn't been cleared successfully\n,the value Infor Cloverleaf(R) Secure Courier %s Client %s still exists"%(self.FinalVersion.strip(".0"),self.CSC_Client_Name))
			self.logFileContent("error.txt")
		else:
			self.printLog("client install place regdit has been cleared successfully")
		#check the service name is removed
		result=os.system('REG QUERY "HKEY_LOCAL_MACHINE\SYSTEM\ControlSet001\Services\CSCClient%s%s">error.txt'%(self.FinalVersion.strip(".0").replace(".",""),self.CSC_Client_Name))
		if result==0:
			self.printLog("client service name hasn't been cleared successfully\n,the value CSCClient%s%s still exists"%(self.FinalVersion.strip(".0").replace(".",""),self.CSC_Client_Name))
			self.logFileContent("error.txt")
		else:
			self.printLog("client service name has been cleared successfully")
		#start menu compare
		if self.checkStartMenu(os.path.join(r"C:\ProgramData\Microsoft\Windows\Start Menu\Programs","Infor Cloverleaf Secure Courier %s Client %s"%(self.FinalVersion.strip(".0"),self.CSC_Client_Name))):
			self.printLog("client start menu Infor Cloverleaf Secure Courier %s Client %s still could be found"%(self.FinalVersion.strip(".0"),self.CSC_Client_Name))
		else:
			self.printLog("client start menu has been removed successfully")
			
	def checkStartMenu(self,appValue):
		if os.path.exists(appValue):
			return True
		else:
			return False
	
	def printLog(self,content):
		print content
		self.logSteps(content)
	
class runtime(object):
	def __init__(self,path):
		self.path=path
		self.properties=Property(os.path.join(path,"installConfig.ini")).parse()
		self.CSC_InstallDir=self.properties['CSC_InstallDir']
		self.CIS_ROOT_Dir=self.properties['CIS_ROOT_Dir']
		self.CSC_Client_Name=self.properties['CSC_Client_Name']
		self.TOOLS_DIR=self.properties['TOOLS_DIR']
		self.REMOTE=self.properties['REMOTE']
		self.LOCAL_IP=self.properties['LOCAL_IP']
		self.PORT=self.properties['PORT']
		self.SuiteBasePath=self.properties['SuiteBasePath']
		self.ratToolPath=self.properties['ratToolPath']
		self.Platform=platform.system()
		
	def server_run(self):
		os.environ['CSC_SERVER_HOME']=os.path.join(self.CSC_InstallDir,"server").replace("\\","/")
		os.environ['TOOLS_DIR']=self.TOOLS_DIR
		# os.environ['HCIROOT']=self.CIS_ROOT_Dir
		if self.Platform!="Windows":
			os.system('export LD_LIBRARY_PATH=$LD_LIBRARY_PATH:%s/tcl/lib'%os.path.join(self.CSC_InstallDir,"server"))
			os.system('export LD_LIBRARY_PATH=$LD_LIBRARY_PATH:%s/bin'%os.path.join(self.CSC_InstallDir,"server"))
		# print "2234"
		os.chdir(self.TOOLS_DIR)
		if self.Platform=="Windows":
			os.system("setroot %s&&perl csctcptest.pl -p 7788>%s"%(self.CIS_ROOT_Dir,os.path.join(self.path,"csc_server.log")))
		else:
			os.system("perl csctcptest.pl -p 7788 -f %s"%os.path.join(self.path,"csc_server.log"))
		# print "3666"
	def client_run(self):
		os.environ['CSC_CLIENT_HOME']=os.path.join(self.CSC_InstallDir,"client_%s"%self.CSC_Client_Name)
		os.environ['REMOTE']=self.REMOTE
		os.environ['LOCAL_IP']=self.LOCAL_IP
		os.environ['PORT']=self.PORT
		os.environ['TOOLS_DIR']=self.TOOLS_DIR
		if self.Platform!="Windows":
			os.environ['DEV_NULL']=r"/dev/null"
			os.system('export LD_LIBRARY_PATH=$LD_LIBRARY_PATH:%s/tcl/lib'%os.path.join(self.CSC_InstallDir,"client_%s"%self.CSC_Client_Name))
			os.system('export LD_LIBRARY_PATH=$LD_LIBRARY_PATH:%s/tcl/lib/tls1.6.7'%os.path.join(self.CSC_InstallDir,"client_%s"%self.CSC_Client_Name))
		
		os.chdir(self.SuiteBasePath)
		if self.Platform=="Windows":
			os.system("setroot %s&&%s sanity>%s"%(self.CIS_ROOT_Dir,self.ratToolPath,os.path.join(self.path,"csc_sanity.log")))
		else:
			os.system("%s sanity>%s"%(self.ratToolPath,os.path.join(self.path,"csc_sanity.log")))
	def running(self):
		if os.path.exists(os.path.join(self.path,"csc_server.log")):
			os.remove(os.path.join(self.path,"csc_server.log"))
		if os.path.exists(os.path.join(self.path,"csc_sanity.log")):
			os.remove(os.path.join(self.path,"csc_sanity.log"))
		p1 = Process(target=self.server_run)
		p2 = Process(target=self.client_run)
		p1.start()
		# print "4444"
		time.sleep(10)
		p2.start()
		p2.join()
		while p2.is_alive():
			pass
		if not p2.is_alive():
			p1.terminate()
			while p1.is_alive():
				pass
		# p1.terminate()
		print('the sanity testing done.')
		
	def windowsTerminateBatch(self):
		win32api.keybd_event(17, 0, 0, 0)
		win32api.keybd_event(67, 0, 0, 0)
		win32api.keybd_event(17, 0, win32con.KEYEVENTF_KEYUP, 0)
		win32api.keybd_event(67, 0, win32con.KEYEVENTF_KEYUP, 0)
		win32api.keybd_event(89, 0, 0, 0)
		win32api.keybd_event(89, 0, win32con.KEYEVENTF_KEYUP, 0)
		win32api.keybd_event(13, 0, 0, 0)
		win32api.keybd_event(13, 0, win32con.KEYEVENTF_KEYUP, 0)
		win32api.keybd_event(13, 0, 0, 0)
		win32api.keybd_event(13, 0, win32con.KEYEVENTF_KEYUP, 0)
	

	
	def checkSanityIssue(self):
		print "Analysising the csc_sanity.log"
		contentList=[]
		with open(os.path.join(self.path,"csc_sanity.log"),"r") as fp:
			for i in fp:
				if i.startswith("unidentified") or i.startswith("unknown"):
					element1=re.search(r"%s/(.*) - (.*)"%self.SuiteBasePath.replace("\\\\","/"),i).group(1)
					element2=re.search(r"%s/(.*) - (.*)"%self.SuiteBasePath.replace("\\\\","/"),i).group(2)
					contentList.append([element1,element2])
		if len(contentList)==0:
			print "there is no sanity issue"
		else:
			print "failure log has been retrived,please refer the below:"
			for i in contentList:
				print i
		return sorted(contentList)
	
class Property(object):
	def __init__(self,fileName):
		self.fileName=fileName
		self.properties={}
		with open(self.fileName,"r") as fp:
			for line in fp:
				line=line.strip()
				if line.find("=")>0 and not line.startswith("#"):
					strs=line.split("=")
					self.properties[strs[0].strip()]=strs[1].strip()
	def parse(self):
		return self.properties
	
	
if __name__ == "__main__":
	Testing="debug"
	
	installObj=silentInstall(Testing)
	if Testing=="debug":
		# installObj.linuxServiceExists()
		installObj.initialRatsEnvironment()
		installObj.rt.running()
		installObj.rt.checkSanityIssue()
		installObj.KillPid("perl.exe")
	else:
		try:
			installObj.runInstall()
		except:
			traceback.print_exc()
		if raw_input("please press enter key to terminate")=="":
			if installObj.Platform=="Windows":
				installObj.KillPid("perl.exe")
			else:
				installObj.KillPid("perl","csctcptest.pl")
			sys.exit(0)
	
	
	
	
	
	
