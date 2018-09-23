#encoding=utf-8
import os
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook import Workbook
import collections


CurrentLocation=os.getcwd()
AffectedFileName="CIS.QPD.CIS6.1.X.Affected_File_List.xlsx"
wb=load_workbook(os.path.join(CurrentLocation,AffectedFileName))
sheet=wb['Windows']
DictVer=collections.OrderedDict()
DictVer['1']=6.1
MaxVerColumn='L'
for i in range(2,(ord(MaxVerColumn)-ord('A'))+1):
	DictVer[str(i)]=sheet.cell(column=i,row=1).value

Column=12
MaxRow=sheet.max_row
os.chdir(CurrentLocation)
try:
	os.mkdir("Expected")
except:
	pass
os.chdir("Expected")

for verIndex in DictVer.keys():
	List=[]
	for rowIndex in range(3,MaxRow+1):
		for columnIndex in range(int(verIndex)+1,(ord(MaxVerColumn)-ord('A'))+2):
			if sheet.cell(column=columnIndex,row=rowIndex).value!=None:
				List.append(sheet.cell(column=1,row=rowIndex).value)
				break
	with open("%s-6.1.4.txt" %DictVer[verIndex],"w") as fp:
		for i in List:
			fp.write(i+'\n')
