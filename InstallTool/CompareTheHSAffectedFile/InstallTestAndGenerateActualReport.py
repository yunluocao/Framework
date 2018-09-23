#encoding=utf-8
import os
import re
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook import Workbook
import collections
import shutil

InstallFolder=r"C:\Users\jwang4\Documents\CIS"
InstallLogDir=r"C:\clover\host\cis6.1\integrator"
DestiInstallFolder=os.getcwd()
ActualFolder="Actual"
try:
	os.mkdir(ActualFolder)
except:
	pass
DestiInstallFolder=os.path.join(DestiInstallFolder,ActualFolder)
FinalVersion="6.1.4"
InstallIssueList=[]

def errorLogJudge(fileName):
	ErrorFlag=False
	with open(fileName,"r") as fp:
		for i in fp:
			if "successfully" in i:
				return ErrorFlag
		else:
			InstallIssueList.append(fileName)
	ErrorFlag=True
	return ErrorFlag


def installMethod(StartVersion,FinalVersion):
	if StartVersion!="6.1":
		print StartVersion
		print os.path.join(InstallFolder,"CIS"+StartVersion.replace(".",""))
		os.chdir(os.path.join(InstallFolder,"CIS"+StartVersion.replace(".","")))
		print os.getcwd()
		os.system("echo Y|install_update.cmd")
	os.chdir(os.path.join(InstallFolder,"CIS"+FinalVersion.replace(".","")))
	os.system("echo Y|install_update.cmd")
	os.chdir(InstallLogDir)
	orginalLogName="install_"+FinalVersion.replace(".","_")+"_0.log"
	InstallNewLogName="install."+FinalVersion+" from "+StartVersion+".log"
	os.rename(orginalLogName,InstallNewLogName)
	if not errorLogJudge(InstallNewLogName):
		os.chdir(os.path.join(InstallFolder,"CIS"+FinalVersion.replace(".","")))
		os.system("echo Y|uninstall_update.cmd")
		os.chdir(InstallLogDir)
		orginalLogName="uninstall_"+FinalVersion.replace(".","_")+"_0.log"
		UninstallNewLogName="uninstall."+FinalVersion+" to "+StartVersion+".log"
		os.rename(orginalLogName,UninstallNewLogName)
		if not errorLogJudge(UninstallNewLogName):
			shutil.copy(InstallNewLogName,DestiInstallFolder)
			os.remove(InstallNewLogName)
			shutil.copy(UninstallNewLogName,DestiInstallFolder)
			os.remove(UninstallNewLogName)



CurrentLocation=r"C:\AInstallation\AffectedFileCheck"
AffectedFileName="CIS.QPD.CIS6.1.X.Affected_File_List.xlsx"
wb=load_workbook(os.path.join(CurrentLocation,AffectedFileName))
sheet=wb['Windows']
VerList=['6.1']

MaxVerColumn='L'
for i in range(2,(ord(MaxVerColumn)-ord('A'))+1):
	VerList.append(str(sheet.cell(column=i,row=1).value))
for i in VerList:
	installMethod(i,FinalVersion)


#uninstall patch, but it would have one potential issue, if the former patch uninstall issue hasn't found, it would results currpt.
for i in VerList[-1:0:-1]:
	os.chdir(os.path.join(InstallFolder,"CIS"+i.replace(".","")))
	os.system("echo Y|uninstall_update.cmd")
print InstallIssueList