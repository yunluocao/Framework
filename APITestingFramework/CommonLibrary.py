#encoding=utf-8
import os
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,colors,borders
import sys
import logging
import time
import collections
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom

reload(sys)
sys.setdefaultencoding('utf8')


class commonTools():
	def __init__(self):
		self.SumTitleLine=['Execution','Desctiption','TestCaseName','APIMethod','Sitename','Api','apiParameter','BinaryFlag','TestDataFile','Status_Code','ExpectedFileName','ExpectedText']
		self.ScriptName="testcase.py"
		self.TestCaseName="Test.xlsx"
		self.initFunction()
		
		
	def initFunction(self):
		'''initial the logger'''
		logger = logging.getLogger(__name__)
		logger.setLevel(level = logging.DEBUG)
		self.logger=logger
		if len(self.logger.handlers)==0:
			if os.path.exists("logbak.txt"):
				os.remove("logbak.txt")
			if os.path.exists("log.txt"):
				os.rename("log.txt","logbak.txt")
			os.system("type nul>log.txt")
			#as at least log once, then the handler would not be 0, otherwise the bak log file would be replaced
			self.logSteps("Start to do the execution")
		
		
	def initialSuite(self,path,suitename):
		if not os.path.exists(os.path.join(path,suitename)):
			os.mkdir(os.path.join(path,suitename))
		os.chdir(os.path.join(path,suitename))
		os.mkdir("files")
		os.mkdir("results")
		if os.path.exists(self.ScriptName) or os.path.exists(self.TestCaseName):
			print "the testcase has already exists, please check it"
			sys.exit(0)
		os.system("type nul>%s"%self.ScriptName)
		wb = Workbook()
		sheet=wb.active
		sheet.title="Summary"
		sheet.append(self.SumTitleLine)
		MaxVerColumn=sheet.max_column
		column_widths=[20]*MaxVerColumn
		for i, column_width in enumerate(column_widths):
			sheet.column_dimensions[get_column_letter(i+1)].width =column_width
		for i in range(1,MaxVerColumn+1):
			sheet.cell(column=i,row=1).border=Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))
			sheet.cell(column=i,row=1).font=Font(name='Calibri',size=11,bold=True)
		wb.save(self.TestCaseName)
		
	def logSteps(self,content):
		'''judge whether need to print in the console, and logger the records in the log file'''
		if len(self.logger.handlers)==0:
			handler = logging.FileHandler("log.txt")
			self.logger.addHandler(handler)
		if content.startswith('\n'):
			self.logger.info('\n'+time.strftime("[%m/%d/%Y %H:%M:%S]",time.localtime())+" "+content.strip())
		else:
			self.logger.info(time.strftime("[%m/%d/%Y %H:%M:%S]",time.localtime())+" "+content)
	
	def transExcelToDict(self,file):
		'''tran the excel data to the dict format'''
		wb=load_workbook(file)
		sheet=wb.active
		TitleList=[]
		MaxVerColumn=sheet.max_column
		DictVer=collections.OrderedDict()
		MaxVerColumn=sheet.max_column
		MaxRow=sheet.max_row
		
		for i in range(1,MaxVerColumn+1):
			if sheet.cell(column=i,row=1).value:
				TitleList.append(str(sheet.cell(column=i,row=1).value))
		for RowNum in range(2,MaxRow+1):
			for index in range(len(TitleList)):
				DictVer[RowNum-2]=collections.OrderedDict(zip(TitleList,[sheet.cell(column=i,row=RowNum).value for i in range(1,MaxVerColumn+1)]))
		return DictVer
		
	def dict_to_xml(self,input_dict, root_tag, node_tag):
		'''define the first level:root_tag, the second level:node_tag
		the third level put the dict key-value,return xml tree'''
		root_name = ET.Element(root_tag)
		for (k, v) in input_dict.items():
			node_name = ET.SubElement(root_name, node_tag)
			for key, val in v.items():
				key = ET.SubElement(node_name, key)
				'''to avoid the error "cannot serialize %r (type %s)" % (text, type(text).__name__), it needs to remove the\n for each value, 
				   otherwise xml won't be parsed'''
				if val:
					key.text =str(val).strip()
					key.text=key.text.replace('\n',"")
				else:
					key.text =val
		return root_name
	def out_xml(self,root,out_file):
		"""Trans the root into xml"""
		rough_string = ET.tostring(root, 'utf-8')
		reared_content = minidom.parseString(rough_string)
		with open(out_file, 'w+') as fs:
			reared_content.writexml(fs, addindent="   ", newl="\n", encoding="utf-8")
		return True
		
	def creat_dict(self,root):
		"""xml trans inot dict：add each point of the tree into the list, and trans the list into dict_init,and accumlate as multiple level dict_new"""
		dict_new = {}
		for key, valu in enumerate(root):
			dict_init = {}
			list_init = []
			for item in valu:
				list_init.append([item.tag, item.text])
				for lists in list_init:
					dict_init[lists[0]] = lists[1]
			dict_new[key] = dict_init
		return dict_new
		
	def read_xml(self,in_path):
		"""read and parse xml file
		in_path: xml path
		return: tree"""
		tree = ET.parse(in_path)
		return tree
		
class Property(object):
	def __init__(self,fileName):
		self.fileName=fileName
		self.properties={}
		with open(self.fileName,"r") as fp:
			for line in fp:
				line=line.strip()
				if line.find("=")>0 and not line.startswith("#"):
					strs=line.split("=")
					self.properties[strs[0].strip()]=strs[1].strip()
	def parse(self):
		return self.properties


if __name__ == "__main__":
	com=commonTools()
	# com.initialSuite(os.getcwd(),"abcd")
	# DictVer=com.transExcelToDict(r"C:\PythonTest\Framework\APITestingFramework\SampleSuite\cbcd\Test.xlsx")
	# root=com.dict_to_xml(DictVer,"xmlRoot","caseRoot")
	# com.out_xml(root,r"C:\PythonTest\Framework\APITestingFramework\Test_Case.xml")
	com.excuteCaseByXML(r"C:\PythonTest\Framework\APITestingFramework\Test_Case.xml")