#encoding=utf-8

from APILibrary import API
from CommonLibrary import commonTools
from CommonLibrary import Property
import os
from openpyxl import load_workbook
import shutil
from remote import telnetAction
from remote import Xfer
import re
from multiprocessing import Process, Pool, Value, Lock, Manager

class RunSuites(object):
	def __init__(self):
		
		self.suiteExcel="Suite.xlsx"
		self.subSuiteExcel="SubSuite.xlsx"
		self.properties=Property("config.ini").parse()
		self.commLib=commonTools()
		print "Do the initial preparation"
		self.commLib.logSteps("Do the initial preparation")
		#eval transform the string dict into the dict
		self.proxies =eval(self.properties['proxies'])
		self.InstallDir=self.properties['HCIROOT']
		self.proxyFlag=eval(self.properties['proxyFlag'])
		self.remoteFlag=eval(self.properties['remoteFlag'])
		if self.remoteFlag:
			self.TestAddress=self.properties['RemoteTestAddress'].strip()
			self.telnet=telnetAction(eval(self.properties['DebugLevel']),self.InstallDir)
		else:
			self.TestAddress=self.properties['LocalTestAddress'].strip()
			self.telnet=None
		print self.TestAddress
		self.api=API(self.telnet,self.proxies,self.TestAddress,self.InstallDir,self.proxyFlag,self.remoteFlag)
		self.xfer=Xfer(eval(self.properties['DebugLevel']))
		self.filesFolder="files"
		self.testFolder="test"
		self.resultsFolder="results"
		self.Dict={}
		print "The initial preparation done"
		self.commLib.logSteps("The initial preparation done")


	def orginalRunSuite(self,path):
		wbSuite=load_workbook(self.suiteExcel)
		sheetSuite=wbSuite.active
		MaxRowSuite=sheetSuite.max_row
		
		for rIndexSuite in range(2,MaxRowSuite+1):
			SuiteName=sheetSuite.cell(column=3,row=rIndexSuite).value
			if sheetSuite.cell(column=1,row=rIndexSuite).value=="Yes":
				if os.path.exists(os.path.join(path,SuiteName,self.subSuiteExcel)):
					wbSubSuite=load_workbook(os.path.join(path,SuiteName,self.subSuiteExcel))
					sheetSubSuite=wbSubSuite.active
					MaxRowSubSuite=sheetSubSuite.max_row
					for rIndexSubSuite in range(2,MaxRowSubSuite+1):
						SubSuiteName=sheetSubSuite.cell(column=3,row=rIndexSubSuite).value
						if sheetSubSuite.cell(column=1,row=rIndexSubSuite).value=="Yes":
							self.initialSite(path,os.path.join(SuiteName,SubSuiteName))
							self.api.readTheExcel(os.path.join(path,SuiteName),SubSuiteName)
						else:
							print SubSuiteName,"is set No Execution Flag"
				else:
					self.initialSite(path,SuiteName)
					self.api.readTheExcel(path,SuiteName)
			else:
				print SuiteName,"is set No Execution Flag"
				
	def runSuites(self,path):
		if os.path.exists(os.path.join(path,self.suiteExcel)):
			SuiteName=path.split("\\")[-1]
			print path,"is running"
			self.commLib.logSteps("%s is running"%path)
			wbSuite=load_workbook(os.path.join(path,self.suiteExcel))
			sheetSuite=wbSuite.active
			MaxRowSuite=sheetSuite.max_row
			for rIndexSuite in range(2,MaxRowSuite+1):
				SuiteName=sheetSuite.cell(column=3,row=rIndexSuite).value
				if sheetSuite.cell(column=1,row=rIndexSuite).value=="Yes":
					path1=os.path.join(path,SuiteName)
					self.runSuites(path1)
				else:
					print SuiteName,"is set No Execution Flag"
					self.commLib.logSteps("%s is set No Execution Flag"%path)
		else:
			SuiteName=path.split("\\")[-1]
			print path,"is running"
			self.commLib.logSteps("%s is running"%path)
			path="\\".join(path.split("\\")[:-1])
			self.initialSite(path,SuiteName)
			self.Dict[SuiteName]={"TotalNum":0,"PassedNum":0,"FailedNum":0,"SkipNum":0}
			self.api.excuteCaseByXML(path,SuiteName,self.Dict[SuiteName])
			print "*"*40
			print "the %s result is"%SuiteName
			print "*"*40
			print "the Total cases's Num:%s"%self.Dict[SuiteName]["TotalNum"]
			print "the Passed cases's Num:%s"%self.Dict[SuiteName]["PassedNum"]
			print "the Failed cases's Num:%s"%self.Dict[SuiteName]["FailedNum"]
			print "the Skip cases's Num:%s\n"%self.Dict[SuiteName]["SkipNum"]
			self.commLib.logSteps("*"*40)
			self.commLib.logSteps("the %s result is"%SuiteName)
			self.commLib.logSteps("*"*40)
			self.commLib.logSteps("the Total cases's Num:%s"%self.Dict[SuiteName]["TotalNum"])
			self.commLib.logSteps("the Passed cases's Num:%s"%self.Dict[SuiteName]["PassedNum"])
			self.commLib.logSteps("the Failed cases's Num:%s"%self.Dict[SuiteName]["FailedNum"])
			self.commLib.logSteps("the Skip cases's Num:%s\n"%self.Dict[SuiteName]["SkipNum"])


	def initialSite(self,path,SuiteGroup):
		siteName=os.listdir(os.path.join(path,SuiteGroup,self.filesFolder))[0]
		if not self.remoteFlag:
			if eval(self.properties['DebugLevel'])==2:
				os.system("hcisiteinit %s"%siteName)
				os.system("xcopy /E/Y %s %s"%(os.path.join(path,SuiteGroup,self.filesFolder,siteName),os.path.join(self.InstallDir,siteName)))
			else:
				os.system("hcisiteinit %s>1.txt"%siteName)
				os.system("xcopy /E/Y %s %s>1.txt"%(os.path.join(path,SuiteGroup,self.filesFolder,siteName),os.path.join(self.InstallDir,siteName)))
		else:
			self.telnet.tn_read_write_Include_Exception("hcisiteinit %s"%siteName)
			self.xfer.setFtpParams('10.86.56.85', 'hci', 'gofish')
			targetDir=self.telnet.tn_returnTargetDir(self.InstallDir,siteName)
			self.xfer.upload(os.path.join(path,SuiteGroup,self.filesFolder,siteName),re.search(r"'(.*)'",targetDir).group(1))
			#don't push the expected file to the remote server
			# self.xfer.upload(os.path.join(path,SuiteGroup,self.resultsFolder),re.search(r"'(.*)'",targetDir).group(1))
	def Summary(self):
		print "*"*40
		print "the Summary result is"
		print "*"*40
		print "the Total cases's Num:%s"%self.api.TotalCase
		print "the Passed cases's Num:%s"%self.api.PassCase
		print "the Failed cases's Num:%s"%self.api.FailedCase
		print "the Skip cases's Num:%s"%self.api.IgnoreCase
		self.commLib.logSteps("*"*40)
		self.commLib.logSteps("the Summary result is")
		self.commLib.logSteps("*"*40)
		self.commLib.logSteps("the Total cases's Num:%s"%self.api.TotalCase)
		self.commLib.logSteps("the Passed cases's Num:%s"%self.api.PassCase)
		self.commLib.logSteps("the Failed cases's Num:%s"%self.api.FailedCase)
		self.commLib.logSteps("the Skip cases's Num:%s"%self.api.IgnoreCase)
		


if __name__=='__main__':
	run=RunSuites()
	# run.telnet.tn_pythonInit(["import requests","import re","import os","import time","from requests.packages.urllib3.exceptions import InsecureRequestWarning","requests.packages.urllib3.disable_warnings(InsecureRequestWarning)","import urllib","from remote import telnetAction"])
	for i in range(2):
		run.runSuites(run.properties['RunSuite'])
	run.Summary()
	if run.remoteFlag:
		run.telnet.exit()
