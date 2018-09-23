import requests
import re
import os
import json
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from CommonLibrary import commonTools
import traceback
#remove the warning message
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
import urllib
import platform
import time
import multiprocessing
from multiprocessing import Process, Pool, Value, Lock, Manager


class API:
	def __init__(self,telnetObject,proxies,TestAddress,InstallDir,proxyFlag=False,remoteFlag=False):
		self.proxyFlag=proxyFlag
		self.proxies = proxies
		self.remoteFlag=remoteFlag
		self.TestAddress=TestAddress
		self.telnet=telnetObject
		self.InstallDir=InstallDir
		if not self.remoteFlag:
			#Don't use the os.system() method, as this process exists only its own process, so use the temporary environment setting.
			os.environ['HCIROOT']=self.InstallDir
		else:
			self.telnet.initEnv(InstallDir)
			''' the above import library has used in the telnet, please import in the below list as well.'''
			self.telnet.tn_pythonInit(["import requests","import re","import os","import time","import platform","from requests.packages.urllib3.exceptions import InsecureRequestWarning","requests.packages.urllib3.disable_warnings(InsecureRequestWarning)","import urllib"])
		self.resultsFolder="results"
		if not self.remoteFlag:
			self.platform=platform.platform()
		else:
			self.platform=self.telnet.tn_read_write_Include_Exception("platform.platform()")
		self.getSessionCSRFTokenID()
		self.TestCaseExcel="Test.xlsx"
		self.TestCaseXML="Test_Case.xml"
		self.resultsFolder="results"
		self.commLib=commonTools()
		self.TotalCase=0
		self.PassCase=0
		self.FailedCase=0
		self.IgnoreCase=0

	def getSessionCSRFTokenID(self):
		headers = {'Authorization': 'Basic YWRtaW5pc3RyYXRvcjpQQHNzd29yZDAx'}
		if not self.proxyFlag:
			response =requests.get('%s/clapi/api/security/csrf'%self.TestAddress, headers=headers, verify=False)
		else:
			response =requests.get('%s/clapi/api/security/csrf'%self.TestAddress, headers=headers, proxies=self.proxies,verify=False)
		sessionID=re.search(r"JSESSIONID=(.*) for",str(response.cookies)).group(1)
		csrfTokenID=re.search(r': "(.*)" ',response.text).group(1)
		self.cookies = {'JSESSIONID':sessionID.strip()}
		self.headers = {
		'Content-Type': 'application/json',
		'Authorization': 'Basic YWRtaW5pc3RyYXRvcjpQQHNzd29yZDAx',
		'X-CSRF-TOKEN': csrfTokenID,
		}

	def POST(self,sitename,api,apiParameter,binaryFlag,testData,status_code,ExpectedFileName,ExpectedText):
		if self.remoteFlag:
			self.PostMethodRemote(sitename,api,apiParameter,binaryFlag,testData,status_code,ExpectedFileName,ExpectedText)
		else:
			self.PostMethodLocal(sitename,api,apiParameter,binaryFlag,testData,status_code,ExpectedFileName,ExpectedText)
	
	def GET(self,api,apiParameter,status_code,ExpectedFileName,ExpectedText):
		if self.remoteFlag:
			GetMethodRemote(api,apiParameter,status_code,ExpectedFileName,ExpectedText)
		else:
			GetMethodLocal(api,apiParameter,status_code,ExpectedFileName,ExpectedText)
			
			
	
	def PostMethodLocal(self,sitename,api,apiParameter,binaryFlag,testData,status_code,ExpectedFileName,ExpectedText):
		dataValue=self.readPostData(binaryFlag,sitename,testData)
		if not self.proxyFlag:
			response = requests.post('%s/clapi/api/%s?%s'%(self.TestAddress,api,apiParameter), headers=self.headers, cookies=self.cookies, data=dataValue, verify=False)
		else:
			response = requests.post('%s/clapi/api/%s?%s'%(self.TestAddress,api,apiParameter), headers=self.headers, cookies=self.cookies, data=dataValue, proxies=self.proxies, verify=False)
		self.resultContentCmp(response.status_code,response.text,status_code,ExpectedFileName,ExpectedText)
	def PostMethodRemote(self,sitename,api,apiParameter,binaryFlag,testData,status_code,ExpectedFileName,ExpectedText):
		self.readPostData(binaryFlag,sitename,testData)
		if not self.proxyFlag:
			self.telnet.tn_read_write_Include_Exception("response = requests.post('%s/clapi/api/%s?%s', headers=%s, cookies=%s, data=dataValue, verify=False)"%(self.TestAddress,api,apiParameter,self.headers,self.cookies))
		else:
			self.telnet.tn_read_write_Include_Exception("response = requests.post('%s/clapi/api/%s?%s', headers=%s, cookies=%s, data=dataValue, proxies=%s, verify=False)"%(self.TestAddress,api,apiParameter,self.headers,self.cookies,self.proxies))
		resStatusCode=self.telnet.tn_read_write_Include_Exception("response.status_code")
		resText=self.telnet.tn_read_write_Include_Exception("response.text")
		self.resultContentCmp(resStatusCode,resText,status_code,ExpectedFileName,ExpectedText)
	
	def readPostData(self,binaryFlag,sitename,testData):
		if not self.remoteFlag:
			if self.platform.startswith("Windows"):
				if not binaryFlag:
					dataValue = open(r'%s\%s\test\%s'%(self.InstallDir,sitename,testData))
				else:
					dataValue = open(r'%s\%s\test\%s'%(self.InstallDir,sitename,testData),'rb').read()
			else:
				if not binaryFlag:
					dataValue = open(r'%s/%s/test/%s'%(self.InstallDir,sitename,testData))
				else:
					dataValue = open(r'%s/%s/test/%s'%(self.InstallDir,sitename,testData),'rb').read()
			return dataValue
		else:
			if self.platform.startswith("Windows"):
				if not binaryFlag:
					self.telnet.tn_read_write_Include_Exception("dataValue = open(r'%s\%s\test\%s')"%(self.InstallDir,sitename,testData))
				else:
					self.telnet.tn_read_write_Include_Exception("dataValue = open(r'%s\%s\test\%s','rb').read()"%(self.InstallDir,sitename,testData))
			else:
				if not binaryFlag:
					self.telnet.tn_read_write_Include_Exception("dataValue = open(r'%s/%s/test/%s')"%(self.InstallDir,sitename,testData))
				else:
					self.telnet.tn_read_write_Include_Exception("dataValue = open(r'%s/%s/test/%s','rb').read()"%(self.InstallDir,sitename,testData))
			


		
	def GetMethodLocal(self,api,apiParameter,status_code,ExpectedFileName,ExpectedText):
		if not self.proxyFlag:
			response = requests.get('%s/clapi/api/%s?%s'%(self.TestAddress,api,apiParameter), headers=self.headers, cookies=self.cookies, verify=False)
		else:
			response = requests.get('%s/clapi/api/%s?%s'%(self.TestAddress,api,apiParameter), headers=self.headers, cookies=self.cookies, proxies=self.proxies, verify=False)
		self.resultContentCmp(response.status_code,response.text,status_code,ExpectedFileName,ExpectedText)
		
		
	def GetMethodRemote(self,api,apiParameter,status_code,ExpectedFileName,ExpectedText):
		if not self.proxyFlag:
			self.telnet.tn_read_write_Include_Exception("response = requests.get('%s/clapi/api/%s?%s'%(self.TestAddress,api,apiParameter), headers=self.headers, cookies=self.cookies, verify=False)\n")
		else:
			self.telnet.tn_read_write_Include_Exception("response = requests.get('%s/clapi/api/%s?%s'%(self.TestAddress,api,apiParameter), headers=self.headers, cookies=self.cookies, proxies=self.proxies, verify=False)\n")
		resStatusCode=self.telnet.tn_read_write_Include_Exception("response.status_code\n")
		resText=self.telnet.tn_read_write_Include_Exception("response.text\n")
		self.resultContentCmp(resStatusCode,resText,status_code,ExpectedFileName,ExpectedText)
	
	
	def resultContentCmp(self,resStatusCode,resText,status_code,ExpectedFileName,ExpectedText):
		if str(resStatusCode).strip()!=str(status_code).strip():
			raise Exception("The resStatusCode and status_code aren't equal\n, and the resStatusCode is %s\n,the status_code is %s\n"%(resStatusCode,status_code))
		if ExpectedFileName:
			
			with open(ExpectedFileName,"rb") as fp:
				content=fp.readline()
			#read the data in the remote server
			# if self.remoteFlag:
				# self.telnet.tn_read_write_Include_Exception("filepath=os.path.join(\"%s\",\"%s\",\"%s\")"%(self.InstallDir,sitename,os.path.basename(ExpectedFileName)))
				# self.telnet.tn_read_write_Include_Exception('fp=open(filepath,"rb")')
				# content=self.telnet.tn_read_write_Include_Exception('fp.readline()')
				
			if not self.platform.startswith("Windows"):
				content=content.replace(r"\r\n",r"\\n")
			if resText.startswith("u'{"):
				resText=re.search(r"u'({.*})'",resText).group(1)
			if content!=resText:
				raise Exception("The content and resText aren't equal\n, and the content is %s\n,the resText is %s\n"%(content,resText))
		else:
			if not self.platform.startswith("Windows"):
				ExpectedText=ExpectedText.replace(r"\r\n",r"\\n")
			if ExpectedText not in resText:
				raise Exception("The ExpectedText not in resText\n, and the ExpectedText is %s\n,the resText is %s\n"%(ExpectedText.replace("\r\n","\\n"),resText))


		
		
		
	def readTheExcel(self,path,suitename):
		wb=load_workbook(os.path.join(path,suitename,self.TestCaseExcel))
		sheet=wb.active
		MaxRow=sheet.max_row
		self.TotalCase+=MaxRow-1
		for rIndex in range(2,MaxRow+1):
			TestCaseName=sheet.cell(column=3,row=rIndex).value
			if sheet.cell(column=1,row=rIndex).value=="Yes":
				Method=sheet.cell(column=4,row=rIndex).value
				sitename=sheet.cell(column=5,row=rIndex).value
				api=sheet.cell(column=6,row=rIndex).value
				apiParameter=sheet.cell(column=7,row=rIndex).value
				#There are many times that you want to send data that is not form-encoded. If you pass in a string instead of a dict, that data will be PostMethod directly.
				#So no need to transform the parameter, as it pass in the string directly.
				apiParameter="&".join([i.split("=")[0]+"="+urllib.quote(i.split("=")[1]) for i in apiParameter.split("&")])
				binaryFlag=sheet.cell(column=8,row=rIndex).value
				testData=sheet.cell(column=9,row=rIndex).value
				status_code=sheet.cell(column=10,row=rIndex).value
				ExpectedFileName=sheet.cell(column=11,row=rIndex).value
				ExpectedFileName=os.path.join(path,suitename,self.resultsFolder,ExpectedFileName)
				print ExpectedFileName
				ExpectedText=sheet.cell(column=12,row=rIndex).value
				if Method.upper()=="POST":
					self.POST(sitename,api,apiParameter,binaryFlag,testData,status_code,ExpectedFileName,ExpectedText)
					print TestCaseName,"is pass"
					self.commLib.logSteps("%s is pass"%TestCaseName)
					self.PassCase+=1
				else:
					self.GET(api,apiParameter,status_code,ExpectedFileName)
					print TestCaseName,"is pass"
					self.commLib.logSteps("%s is pass"%TestCaseName)
					self.PassCase+=1
			else:
				print TestCaseName,"is set No Execution Flag"
				self.commLib.logSteps("%s is set No Execution Flag"%TestCaseName)
				self.IgnoreCase+=1
					
	def excuteCaseByXML(self,path,suitename,DictSuiteNumValue):
		tree=self.commLib.read_xml(os.path.join(path,suitename,self.TestCaseXML))
		dict_new=self.commLib.creat_dict(tree.getroot())
		self.TotalCase+=len(dict_new.items())
		DictSuiteNumValue["TotalNum"]=len(dict_new.items())
		for (k, v) in dict_new.items():
			TestCaseName=v["TestCaseName"]
			if v["Execution"]=="Yes":
				paramter=""
				funName=v["APIMethod"].strip().upper()
				for key, val in v.items():
					if key not in ["Execution","TestCaseName","Desctiption","APIMethod"]:
						if not val:
							val=""
						if key=="apiParameter":
							val="&".join([i.split("=")[0]+"="+urllib.quote(i.split("=")[1]) for i in val.split("&")])
						elif key=="ExpectedFileName":
							val=os.path.join(path,suitename,self.resultsFolder,val)
						# this for the address translate
						if key=="ExpectedFileName":
							paramter+=key+"="+"r"+"\""+str(val)+"\""+","
						else:
							paramter+=key+"="+"\""+str(val)+"\""+","
				paramter=paramter.strip(",")
				command="self."+funName+"(%s)"%paramter
				try:
					eval(command)
					print TestCaseName,"is pass"
					self.commLib.logSteps("%s is pass"%TestCaseName)
					self.PassCase+=1
					DictSuiteNumValue["PassedNum"]+=1
				except Exception as e:
					print TestCaseName,"is failed"
					self.commLib.logSteps("%s is failed"%TestCaseName)
					#exception print
					traceback.print_exc()
					#exception logger
					self.commLib.logger.exception("Exception Logged")
					self.FailedCase+=1
					DictSuiteNumValue["FailedNum"]+=1
			else:
				print TestCaseName,"is set No Execution Flag"
				self.commLib.logSteps("%s is set No Execution Flag"%TestCaseName)
				self.IgnoreCase+=1
				DictSuiteNumValue["SkipNum"]+=1
	
	
if __name__=='__main__':
	api=API()
	# api.readTheExcel(os.getcwd(),"abcd")
	s='api.PostMethod(status_code="200",api="ar17522command/test/ldl",sitename="ar17522",apiParameter="n=&i=&e=UTF-8&f=eof&V=2014&v=test&M=C_MOVE&c=%7BTransferSyntax%201.2.840.10008.1.2%7D",ExpectedFileName="C:\PythonTest\Framework\APITestingFramework\SampleSuite\cbcd\results\expected3.txt",ExpectedText="",binaryFlag="True",testData="ldltest.dat")'
	eval(s)


